"""
core_logic.py — 分單系統 V4（超商版）
v4 改動：
  1. 完全砍掉 copy.deepcopy(reader.pages[pi])
     改用 mini PDF 隔離法：每頁獨立 PdfWriter → BytesIO → PdfReader → 取 pages[0]
     避開 Windows 上 deepcopy 的記憶體 leak（18 GB → 預期 < 5 GB）
  2. generate_pdfs 接受 pdf_paths（list），多份 PDF 合併後統一分類輸出
     呼叫端傳 [path1, path2, ...] 即可；oid → page_indices 以全域頁碼計算
  3. 保留 v3 所有優化：_SCAN_CACHE、page.chars、gc.collect、clear_scan_cache
  4. 每 50 頁做一次 gc.collect（Windows 記憶體回收保險）
  5. 分類間 gc.collect（與 v3 一致）
"""

import os, re, gc, platform
from pathlib import Path
from io import BytesIO
from collections import defaultdict

from PIL import Image, ImageDraw, ImageFont
from pypdf import PdfReader, PdfWriter
import pdfplumber
import openpyxl

# ══════════════════════════════════════════════════════════════════════════════
#  字型路徑
# ══════════════════════════════════════════════════════════════════════════════
if platform.system() == "Windows":
    _FONT_CANDIDATES = [
        r"C:\Windows\Fonts\msjhbd.ttc",
        r"C:\Windows\Fonts\msjh.ttc",
        r"C:\Windows\Fonts\mingliu.ttc",
    ]
else:
    _FONT_CANDIDATES = [
        "/usr/share/fonts/opentype/noto/NotoSansCJK-Bold.ttc",
        "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc",
        "/usr/share/fonts/truetype/noto/NotoSansCJK-Bold.ttc",
        "/usr/share/fonts/truetype/wqy/wqy-zenhei.ttc",
        "/usr/share/fonts/truetype/wqy/wqy-microhei.ttc",
    ]

FONT_PATH = next((p for p in _FONT_CANDIDATES if os.path.exists(p)), None)

if FONT_PATH is None and platform.system() != "Windows":
    try:
        import urllib.request
        _dl_path = str(Path(__file__).parent.parent / "NotoSansCJK-Bold.ttc")
        if not os.path.exists(_dl_path):
            print("下載中文字型...")
            urllib.request.urlretrieve(
                "https://github.com/googlefonts/noto-cjk/raw/main/Sans/OTC/NotoSansCJK-Bold.ttc",
                _dl_path
            )
        FONT_PATH = _dl_path
        print(f"字型已載入：{_dl_path}")
    except Exception as e:
        print(f"字型下載失敗：{e}")

# ══════════════════════════════════════════════════════════════════════════════
#  注入設定（由 main.py 在每次處理前呼叫）
# ══════════════════════════════════════════════════════════════════════════════
_SPEC_MAP: dict = {}
_CLASSIFICATION_RULES: dict = {}
_GIFT_RULES: dict = {}
_DISCOUNT_EXCLUDES: list = []
_REMARK_EXCLUDES: list = []
_PURCHASE_MAP: dict = {}
_BOX_TYPES: list = []
_DISCOUNT_THRESHOLD: int = 0
_DISCOUNT_MIN_COUNT: int = 0
_AVG_PRICE_THRESHOLD: int = 450
# {child名稱: 縮寫}，由 categories 的 child_abbr 欄位注入
_CHILD_ABBR: dict = {}
# 子分類統計輸出順序（寫死）：每個 list 是一行
# 第一行：洗髮精 + 潤髮乳
# 第二行：沐浴乳 + 身體乳 + 幕斯面膜
# 第三行：精粹
_SUMMARY_ORDER: list[list[str]] = [
    ['C', 'F', 'C潤', 'F潤'],
    ['抗痘沐', '水光沐', '痘乳', '白乳', '早C', '晚A'],
    ['橘', '綠'],
]


def inject_spec_map(sm: dict):
    global _SPEC_MAP
    _SPEC_MAP = sm or {}


def inject_classification_rules(rules: dict):
    global _CLASSIFICATION_RULES
    _CLASSIFICATION_RULES = rules or {}


def inject_gift_rules(rules: dict):
    global _GIFT_RULES
    _GIFT_RULES = rules or {}


def inject_discount_excludes(values):
    global _DISCOUNT_EXCLUDES
    result = []
    for v in values:
        try:
            result.append(int(str(v).strip()))
        except (ValueError, TypeError):
            pass
    _DISCOUNT_EXCLUDES = result


def inject_remark_excludes(values):
    global _REMARK_EXCLUDES
    _REMARK_EXCLUDES = [str(v).strip() for v in values if str(v).strip()]


def inject_purchase_map(pm: dict):
    global _PURCHASE_MAP
    _PURCHASE_MAP = {str(k).strip(): int(v) for k, v in pm.items() if k and v}


def inject_box_types(bt: list):
    global _BOX_TYPES
    _BOX_TYPES = bt or []


def inject_discount_threshold(v):
    global _DISCOUNT_THRESHOLD
    try:
        _DISCOUNT_THRESHOLD = int(v) if v not in (None, '', 'None') else 0
    except (ValueError, TypeError):
        _DISCOUNT_THRESHOLD = 0


def inject_discount_min_count(v):
    global _DISCOUNT_MIN_COUNT
    try:
        _DISCOUNT_MIN_COUNT = int(v) if v not in (None, '', 'None') else 0
    except (ValueError, TypeError):
        _DISCOUNT_MIN_COUNT = 0


def inject_avg_price_threshold(v):
    global _AVG_PRICE_THRESHOLD
    try:
        _AVG_PRICE_THRESHOLD = int(v) if v not in (None, '', 'None') else 450
    except (ValueError, TypeError):
        _AVG_PRICE_THRESHOLD = 450

def inject_child_abbr(categories: dict):
    """從 categories dict 中抽出每個子分類的 abbr，建立 {child_name: abbr} 對照表"""
    global _CHILD_ABBR
    result = {}
    for parent, children in (categories or {}).items():
        if isinstance(children, dict):
            for child_name, child_info in children.items():
                if isinstance(child_info, dict) and child_info.get('abbr'):
                    result[child_name] = str(child_info['abbr']).strip()
                elif isinstance(child_info, str) and child_info.strip():
                    result[child_name] = child_info.strip()
        elif isinstance(children, list):
            for item in children:
                if isinstance(item, dict):
                    name = item.get('name', '')
                    abbr = item.get('abbr', '')
                    if name and abbr:
                        result[name] = str(abbr).strip()
    _CHILD_ABBR = result


# ══════════════════════════════════════════════════════════════════════════════
#  性別判斷
# ══════════════════════════════════════════════════════════════════════════════
M_CH = set(
    '豪傑偉強勇剛宏峰毅威哲明志仁義德廷秉揚博賢'
    '文建俊冠家國宇承彥柏翔凱嘉智信忠'
    '成昌興隆元中正立達良榮慶福輝勳銘'
    '睿軒宸鈞皓晨洋昱昊昇晉晟景曜程祐佑'
    '書彬章維綸緯紹遠迪進運通順豐茂萬全世民'
    '岩岳山川海濤澤浩瀚源潤霖森松林樺楷楠武'
    '健勝雄震雷士仕丞仲伯均其奇啟希學孝'
    '寬富寶宣宗守實弘彰征御懷敬慎新旭星'
    '春冬夏禹舜堯一少永江漢天光瑞祥'
    '龍虎鴻麟聖高南北東西京展超卓'
    '群逸翰聰霆磊鵬朗則言'
)


def gender(name: str) -> str:
    name = str(name).strip()
    for ch in name:
        if ch in M_CH:
            return 'male'
    return 'female'


# ══════════════════════════════════════════════════════════════════════════════
#  Excel 解析（與 v3 相同，支援多檔合併）
# ══════════════════════════════════════════════════════════════════════════════
def parse_excel(path: str) -> list:
    """解析單一 Excel 檔，回傳訂單列表"""
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb['Order'] if 'Order' in wb.sheetnames else wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    header = [str(c).strip() if c is not None else '' for c in rows[0]]

    def col(name, row):
        try:
            idx = header.index(name)
            v = row[idx]
            return v if v is not None else ''
        except (ValueError, IndexError):
            return ''

    order_data = {}

    for row in rows[1:]:
        oid = str(col('訂單編號', row)).strip()
        if not oid or oid in ('None', ''):
            continue

        spec_raw = str(col('商品規格', row)).strip()
        qty_raw  = col('數量', row)
        try:
            qty = max(1, int(float(str(qty_raw)))) if qty_raw != '' else 1
        except (ValueError, TypeError):
            qty = 1

        # 訂購人姓名（用於 VIP 等級、歷史購買查詢、排序）
        buyer    = str(col('購買人姓名', row) or '').strip()
        if buyer in ('None', ''):
            buyer = ''
        # 收貨人姓名（訂單卡顯示、物流系統）
        recipient = str(col('姓名', row) or col('收件者姓名', row) or '').strip()
        if recipient in ('None', ''):
            recipient = ''
        # 收貨人手機
        phone_raw = str(col('手機', row) or col('收件人電話', row) or '').strip()
        if phone_raw in ('None', ''):
            phone_raw = ''
        # 物流選擇（"7-11 取貨付款" / "7-11 純取貨" / "宅配..."）
        shipping_method = str(col('物流選擇', row) or '').strip()
        if shipping_method in ('None', ''):
            shipping_method = ''

        # name 欄保留為「訂購人」優先、收貨人 fallback（與舊行為相容）
        name    = buyer or recipient
        account = str(col('會員帳號', row) or col('Email', row) or '').strip()
        remark_raw = str(col('買家備註', row) or col('訂單備註', row) or '').strip()
        remark = '' if remark_raw.startswith('/') else remark_raw

        def safe_int(v):
            try:
                return int(float(str(v))) if v not in (None, '', 'None') else 0
            except (ValueError, TypeError):
                return 0

        shipping     = safe_int(col('運費', row))
        discount     = safe_int(col('總折扣', row))
        order_total  = safe_int(col('訂單總金額', row))
        store_credit = safe_int(col('購物金使用', row))
        vip_level_raw = str(col('會員等級', row) or '').strip()
        if vip_level_raw in ('None', ''):
            vip_level_raw = ''

        # 訂購日期 = 「建立時間」欄；datetime → 'YYYY/MM/DD HH:MM:SS'
        order_time_raw = col('建立時間', row)
        if hasattr(order_time_raw, 'strftime'):
            order_time = order_time_raw.strftime('%Y/%m/%d %H:%M:%S')
        elif order_time_raw not in (None, '', 'None'):
            s = str(order_time_raw).strip()
            order_time = s.replace('-', '/', 2) if len(s) >= 10 and s[4] == '-' else s
        else:
            order_time = ''

        if oid not in order_data:
            order_data[oid] = {
                'oid':          oid,
                'name':         name if name not in ('None', '') else '',
                'buyer':        buyer,
                'recipient':    recipient,
                'phone':        phone_raw,
                'shipping_method': shipping_method,
                'account':      account[:8].lower() if account not in ('None', '') else '',
                'remark':       remark if remark not in ('None', '') else '',
                'specs':        [],
                'items':        [],
                'shipping':     shipping,
                'discount':     discount,
                'order_total':  order_total,
                'store_credit': store_credit,
                'vip_level':    vip_level_raw,
                'order_time':   order_time,
            }
        else:
            if shipping > 0:
                order_data[oid]['shipping'] = shipping
            if discount > 0:
                order_data[oid]['discount'] = discount
            if order_total > 0 and order_data[oid].get('order_total', 0) <= 0:
                order_data[oid]['order_total'] = order_total
            if store_credit > 0 and order_data[oid].get('store_credit', 0) <= 0:
                order_data[oid]['store_credit'] = store_credit
            if vip_level_raw and not order_data[oid].get('vip_level'):
                order_data[oid]['vip_level'] = vip_level_raw
            if name and name not in ('None', '') and not order_data[oid]['name']:
                order_data[oid]['name'] = name
            if remark and remark not in ('None', '') and not order_data[oid]['remark']:
                order_data[oid]['remark'] = remark

        if spec_raw and spec_raw not in ('None', ''):
            order_data[oid]['items'].append({'spec': spec_raw, 'qty': qty})
            for _ in range(qty):
                order_data[oid]['specs'].append(spec_raw)

    orders = []
    for oid, data in order_data.items():
        data['effective_discount'] = data['discount'] - data['shipping']
        account = data.get('account', '')
        name    = data.get('name', '')
        key     = f"{account.lower()}:{name}" if account else ''
        data['purchase_count'] = _PURCHASE_MAP.get(key, 0) if key else 0

        total_qty = 0
        for it in data.get('items', []):
            spec     = it.get('spec', '')
            row_qty  = it.get('qty', 1)
            info     = _SPEC_MAP.get(spec)
            if info:
                conds    = info if isinstance(info, list) else [info]
                unit     = sum(c.get('qty', 1) for c in conds)
                total_qty += unit * row_qty
            else:
                total_qty += row_qty
        data['total_qty'] = total_qty

        order_total = data.get('order_total', 0)
        if order_total > 0 and total_qty > 0:
            data['avg_price'] = order_total / total_qty
            data['is_low_avg'] = data['avg_price'] < _AVG_PRICE_THRESHOLD
        else:
            data['avg_price']  = 0
            data['is_low_avg'] = False

        orders.append(data)

    return orders


def parse_excels(paths: list) -> list:
    """
    合併多份 Excel，回傳合併後的訂單列表。
    同一訂單號出現在多份 Excel 時，以第一次出現的資料為主，後續只補 items。
    """
    merged = {}
    for path in paths:
        orders = parse_excel(path)
        for o in orders:
            oid = o['oid']
            if oid not in merged:
                merged[oid] = o
            else:
                # 補充 items（不重複）
                existing_specs = {(it['spec'], it['qty']) for it in merged[oid]['items']}
                for it in o['items']:
                    key = (it['spec'], it['qty'])
                    if key not in existing_specs:
                        merged[oid]['items'].append(it)
                        existing_specs.add(key)
                        for _ in range(it['qty']):
                            merged[oid]['specs'].append(it['spec'])
    return list(merged.values())


# ══════════════════════════════════════════════════════════════════════════════
#  核心分類邏輯（與 v3 相同）
# ══════════════════════════════════════════════════════════════════════════════
def classify_order(order: dict) -> tuple:
    remark       = (order.get('remark') or '').strip()
    store_credit = order.get('store_credit', 0)
    items        = order.get('items', [])

    has_remark = bool(remark and remark not in ('None', '') and remark not in _REMARK_EXCLUDES)
    is_high_discount = False
    if _DISCOUNT_THRESHOLD > 0 and store_credit > _DISCOUNT_THRESHOLD:
        if store_credit not in set(_DISCOUNT_EXCLUDES):
            is_high_discount = True

    order['has_remark']       = has_remark
    order['is_high_discount'] = is_high_discount

    is_tail = False

    if not items:
        return '特殊單', True, is_tail

    counts = defaultdict(int)
    for item in items:
        spec = item['spec']
        qty  = item.get('qty', 1)
        info = _SPEC_MAP.get(spec)
        if not info:
            return '特殊單', True, is_tail
        conds = info if isinstance(info, list) else [info]
        for cond in conds:
            parent = cond.get('parent', '')
            child  = cond.get('child', '')
            unit   = cond.get('qty', 1)
            counts[(parent, child)] += unit * qty

    if not counts:
        return '特殊單', True, is_tail

    matched_cat = _match_rules(counts)
    if matched_cat:
        return matched_cat, False, is_tail

    return '特殊單', True, is_tail


def _match_rules(counts: dict) -> str | None:
    for cat_name, conditions in _CLASSIFICATION_RULES.items():
        if _rule_matches(counts, conditions):
            return cat_name
    return None


def _rule_matches(counts: dict, conditions: list) -> bool:
    exact_conds  = [c for c in conditions if c.get('child', '')]
    parent_conds = [c for c in conditions if not c.get('child', '')]

    rule_parents_exact  = {c['parent'] for c in exact_conds}
    rule_parents_parent = {c['parent'] for c in parent_conds}
    all_rule_parents    = rule_parents_exact | rule_parents_parent

    order_parents = {p for (p, c) in counts.keys()}

    if order_parents != all_rule_parents:
        return False

    if exact_conds:
        order_sig = frozenset((p, c, n) for (p, c), n in counts.items()
                              if p in rule_parents_exact)
        rule_sig  = frozenset(
            (cond['parent'], cond['child'], cond.get('qty', 1))
            for cond in exact_conds
        )
        if order_sig != rule_sig:
            return False

    for cond in parent_conds:
        parent   = cond['parent']
        expected = cond.get('qty', 1)
        actual   = sum(n for (p, c), n in counts.items() if p == parent)
        if actual != expected:
            return False

    return True


# ══════════════════════════════════════════════════════════════════════════════
#  贈品查詢
# ══════════════════════════════════════════════════════════════════════════════
def get_gift(cat: str, sex: str = 'female') -> str:
    rule = _GIFT_RULES.get(cat, {})
    if not rule:
        return f'（找不到贈品規則：{cat}）'
    return rule.get(sex) or rule.get('female') or '請人工確認'


# ══════════════════════════════════════════════════════════════════════════════
#  特殊分類抽出（備註訂單 / 折扣單）
# ══════════════════════════════════════════════════════════════════════════════
REMARK_CAT   = '備註訂單'
DISCOUNT_CAT = '折扣單'
SPECIAL_CAT  = '特殊單'


def reassign_special_categories(orders: list) -> None:
    for o in orders:
        if o.get('has_remark'):
            o['cat']     = REMARK_CAT
            o['special'] = True

    if _DISCOUNT_THRESHOLD <= 0 or _DISCOUNT_MIN_COUNT <= 0:
        return

    by_cat = defaultdict(list)
    for o in orders:
        if o.get('cat') == REMARK_CAT:
            continue
        by_cat[o.get('cat')].append(o)

    for cat, lst in by_cat.items():
        high_discount_orders = [o for o in lst if o.get('is_high_discount')]
        if len(high_discount_orders) >= _DISCOUNT_MIN_COUNT:
            for o in high_discount_orders:
                o['cat']     = DISCOUNT_CAT
                o['special'] = True


def _vip_rank(level: str) -> int:
    s = (level or '').strip().upper()
    if s == 'SVIP':
        return 2
    if s == 'VIP':
        return 1
    return 0


def _layer_rank(o: dict) -> tuple:
    pc      = o.get('purchase_count', 0)
    high_d  = o.get('is_high_discount', False)
    low_avg = o.get('is_low_avg', False)
    vip     = _vip_rank(o.get('vip_level', ''))

    if pc <= 0:
        return (1, vip, o.get('oid', ''))
    if high_d:
        return (6, vip, o.get('oid', ''))
    if low_avg:
        return (5, vip, o.get('oid', ''))
    return (2 + vip, 0, o.get('oid', ''))


def _special_group_key(o: dict) -> str:
    items = o.get('items', [])
    norm  = sorted([(it.get('spec', ''), it.get('qty', 1)) for it in items])
    return '|'.join(f'{s}*{q}' for s, q in norm)


def sort_orders_in_category(cat: str, lst: list) -> list:
    if cat == SPECIAL_CAT:
        groups = defaultdict(list)
        for o in lst:
            groups[_special_group_key(o)].append(o)
        sorted_keys = sorted(
            groups.keys(),
            key=lambda k: (
                -len(groups[k]),
                -sum(o.get('total_qty', 0) for o in groups[k]),
                k
            )
        )
        result = []
        for k in sorted_keys:
            result.extend(sorted(groups[k], key=_layer_rank))
        return result

    if cat in (REMARK_CAT, DISCOUNT_CAT):
        groups = defaultdict(list)
        for o in lst:
            groups[_special_group_key(o)].append(o)
        sorted_keys = sorted(
            groups.keys(),
            key=lambda k: (-len(groups[k]), k)
        )
        result = []
        for k in sorted_keys:
            result.extend(sorted(groups[k], key=_layer_rank))
        return result

    return sorted(lst, key=_layer_rank)


# ══════════════════════════════════════════════════════════════════════════════
#  PDF 頁碼定位（v3 記憶體優化版，支援多份 PDF）
# ══════════════════════════════════════════════════════════════════════════════
_SCAN_CACHE = {}  # {cache_key: (oid_pages, order_total_pos)}


def _scan_pdf_full(pdf_path: str) -> tuple[dict, dict]:
    """
    一次掃描完整 PDF，同時取得：
      - oid_pages: {oid: [page_indices]}  （頁碼為該 PDF 內的索引）
      - order_total_pos: {oid: (page_idx_in_order, y_pdf_bottom, logistics_y)}
        其中 logistics_y 為跨頁訂單第一頁「物流編號」（訂單資訊區，x0<100）的 y，
        單頁時為 None。
    """
    oid_pages       = {}
    order_total_pos = {}

    with pdfplumber.open(pdf_path) as pdf:
        total = len(pdf.pages)
        page_info = []

        for idx in range(total):
            page = pdf.pages[idx]
            ph   = page.height

            try:
                chars = page.chars
            except Exception:
                chars = []

            text = ''.join(c.get('text', '') for c in chars)

            # OID：# 開頭 7+ 字母數字
            m   = re.search(r'#([A-Z0-9]{7,})(?![A-Z0-9])', text)
            oid = m.group(1) if m else None

            has_content = bool(re.search(r'總計|訂單金額|訂單⾦額|\$[\d,]+', text))

            # 找「訂單金額」y 座標（滑動視窗）
            total_y = None
            n = len(chars)
            for i in range(n - 3):
                c0 = chars[i].get('text', '')
                c1 = chars[i + 1].get('text', '')
                c2 = chars[i + 2].get('text', '')
                c3 = chars[i + 3].get('text', '')
                if c0 == '訂' and c1 == '單' and (c2 == '金' or c2 == '⾦') and c3 == '額':
                    bottom  = chars[i].get('bottom', 0)
                    total_y = ph - bottom
                    break

            # 找「總計」y 座標（優先使用此值，比訂單金額更精準）
            subtotal_y = None
            for i in range(n - 1):
                if chars[i].get('text', '') == '總' and chars[i + 1].get('text', '') == '計':
                    bottom     = chars[i].get('bottom', 0)
                    subtotal_y = ph - bottom
                    break

            # 找訂單資訊區「物流編號」y（x0 < 100，排除超商面單區）
            logistics_y = None
            for i in range(n - 3):
                t0 = chars[i].get('text', ''); t1 = chars[i+1].get('text', '')
                t2 = chars[i+2].get('text', ''); t3 = chars[i+3].get('text', '')
                if t0 == '物' and t1 == '流' and t2 == '編' and t3 == '號':
                    x0 = chars[i].get('x0', 999)
                    if x0 < 100:
                        logistics_y = ph - chars[i].get('bottom', 0)
                        break

            # 「總計」上緣 y（單頁統計標籤 Y 起點）
            subtotal_top_y = None
            for i in range(n - 1):
                if chars[i].get('text', '') == '總' and chars[i+1].get('text', '') == '計':
                    subtotal_top_y = ph - chars[i].get('top', 0)
                    break

            # 「總計」右緣 x（單頁統計標籤左緣）
            subtotal_x1 = None
            for i in range(n - 1):
                if chars[i].get('text', '') == '總' and chars[i+1].get('text', '') == '計':
                    subtotal_x1 = chars[i+1].get('x1', 0)
                    break

            # 「商品編號」x0（跨頁統計標籤左緣，x0>100 排除序號欄）
            spec_col_x0 = None
            for i in range(n - 3):
                t0=chars[i].get('text',''); t1=chars[i+1].get('text','')
                t2=chars[i+2].get('text',''); t3=chars[i+3].get('text','')
                if t0=='商' and t1=='品' and t2=='編' and t3=='號':
                    x0 = chars[i].get('x0', 0)
                    if x0 > 100:
                        spec_col_x0 = x0
                        break

            # 「小計」右緣 x（統計標籤右緣）
            subtotal_x_right = None
            for i in range(n - 1):
                if chars[i].get('text', '') in ('小', '⼩') and chars[i+1].get('text', '') == '計':
                    subtotal_x_right = chars[i+1].get('x1', 0)
                    break

            # 收件人電話下緣 y（跨頁統計標籤 Y，x0<100 確保是訂單資訊區）
            phone_bot_y = None
            for i in range(n - 1):
                if chars[i].get('text', '') == '電' and chars[i+1].get('text', '') == '話':
                    if chars[i].get('x0', 999) < 100:
                        phone_bot_y = ph - chars[i+1].get('bottom', 0)
                        break

            page_info.append({
                'oid':              oid,
                'has_content':      has_content,
                'total_y':          total_y,
                'subtotal_y':       subtotal_y,
                'subtotal_top_y':   subtotal_top_y,
                'subtotal_x1':      subtotal_x1,
                'spec_col_x0':      spec_col_x0,
                'subtotal_x_right': subtotal_x_right,
                'logistics_y':      logistics_y,
                'phone_bot_y':      phone_bot_y,
            })

            try:
                page.flush_cache()
            except Exception:
                pass

        # Pass 2: 拼出 oid_pages（含跨頁）
        i = 0
        while i < total:
            info = page_info[i]
            if not info['oid']:
                i += 1
                continue
            oid          = info['oid']
            page_indices = [i]
            j = i + 1
            while j < total:
                nxt = page_info[j]
                if (not nxt['oid']) and nxt['has_content']:
                    page_indices.append(j)
                    j += 1
                else:
                    break
            oid_pages[oid] = page_indices
            i = j

        # Pass 3: 算出 order_total_pos
        # 9-tuple: (page_idx, total_y, logi_y, subtotal_y,
        #           subtotal_top_y, subtotal_x1, spec_col_x0, subtotal_x_right, phone_bot_y)
        # 修正：跨頁訂單各欄位可能散落在不同頁，需掃全部頁面分別取得
        for oid, indices in oid_pages.items():
            is_multipage = (len(indices) > 1)
            first_pi = indices[0]
            fi = page_info[first_pi]

            # 從所有頁面收集各欄位（取最後找到的 total_y，最早找到的 subtotal）
            total_y         = None
            subtotal_y      = None
            subtotal_top_y  = None
            subtotal_x1     = None
            spec_col_x0     = None
            subtotal_x_right= None
            page_idx_in_order = 0

            for i_in_order, pi in enumerate(indices):
                if pi >= total:
                    continue
                info = page_info[pi]
                # total_y / subtotal_y：以最後一頁找到的為準（訂單金額在最後頁）
                if info['total_y'] is not None:
                    total_y = info['total_y']
                    page_idx_in_order = i_in_order
                if info['subtotal_y'] is not None:
                    subtotal_y = info['subtotal_y']
                # subtotal_top_y / x1 / spec_col_x0 / subtotal_x_right：取第一個找到的
                if subtotal_top_y is None and info['subtotal_top_y'] is not None:
                    subtotal_top_y = info['subtotal_top_y']
                if subtotal_x1 is None and info['subtotal_x1'] is not None:
                    subtotal_x1 = info['subtotal_x1']
                if spec_col_x0 is None and info['spec_col_x0'] is not None:
                    spec_col_x0 = info['spec_col_x0']
                if subtotal_x_right is None and info['subtotal_x_right'] is not None:
                    subtotal_x_right = info['subtotal_x_right']

            logi_y  = fi['logistics_y'] if is_multipage else None
            phone_y = fi['phone_bot_y'] if is_multipage else None

            if total_y is not None or subtotal_y is not None:
                order_total_pos[oid] = (
                    page_idx_in_order,
                    total_y or subtotal_y or 24,
                    logi_y,
                    subtotal_y,
                    subtotal_top_y,
                    subtotal_x1,
                    spec_col_x0,
                    subtotal_x_right,
                    phone_y,
                )
            else:
                order_total_pos[oid] = (0, 24, None, None, None, None, None, None, None)


        del page_info

    gc.collect()
    return oid_pages, order_total_pos


def _cache_key(pdf_path: str) -> str:
    try:
        st = os.stat(pdf_path)
        return f'{pdf_path}|{st.st_mtime_ns}|{st.st_size}'
    except Exception:
        return pdf_path


def _get_or_scan(pdf_path: str) -> tuple[dict, dict]:
    key = _cache_key(pdf_path)
    if key in _SCAN_CACHE:
        return _SCAN_CACHE[key]
    result = _scan_pdf_full(pdf_path)
    _SCAN_CACHE.clear()
    _SCAN_CACHE[key] = result
    return result


def locate_order_total_position(pdf_path: str, oid_pages: dict) -> dict:
    _, order_total_pos = _get_or_scan(pdf_path)
    if oid_pages:
        return {oid: order_total_pos.get(oid, (0, 24, None, None, None, None, None, None, None)) for oid in oid_pages.keys()}
    return order_total_pos


def locate_pdf_pages(pdf_path: str) -> dict:
    oid_pages, _ = _get_or_scan(pdf_path)
    return oid_pages


def clear_scan_cache():
    _SCAN_CACHE.clear()
    gc.collect()


# ══════════════════════════════════════════════════════════════════════════════
#  v4 核心：mini PDF 隔離法（取代 deepcopy）
# ══════════════════════════════════════════════════════════════════════════════
def _isolated_page(reader: PdfReader, page_index: int):
    """
    用 mini PDF 隔離法取出第 page_index 頁，與原 reader 完全獨立。
    完全取代 copy.deepcopy(reader.pages[page_index])。

    原理：
      1. 建一個暫時的 PdfWriter，只放這一頁
      2. 寫入 BytesIO（在記憶體，不碰磁碟）
      3. 用新的 PdfReader 讀這個 mini PDF
      4. 回傳 mini_reader.pages[0]

    優點：
      - 與原 reader 的 indirect objects 完全無關，蓋章不會污染其他頁
      - 避開 Windows 上 deepcopy 的記憶體 leak
    """
    mini_w = PdfWriter()
    mini_w.add_page(reader.pages[page_index])
    buf = BytesIO()
    mini_w.write(buf)
    buf.seek(0)
    mini_r = PdfReader(buf)
    page   = mini_r.pages[0]
    # mini_w / buf / mini_r 留在 caller 清理（caller del 完後 gc 回收）
    return page, mini_r, buf


# ══════════════════════════════════════════════════════════════════════════════
#  PDF 輸出（v4：多份 PDF 合併 + mini PDF 隔離法）
# ══════════════════════════════════════════════════════════════════════════════
def generate_pdfs(orders: list, pdf_paths, out_dir: str, print_gift: bool,
                  order_total_pos: dict = None) -> list:
    """
    orders          : 已分類排序的訂單列表，每筆含 page_indices（全域頁碼）
    pdf_paths       : str（單份）或 list[str]（多份，v4 新增）
                      多份時，全域頁碼 = 第 0 份頁碼 → 第 1 份接續 → ...
    out_dir         : 輸出資料夾
    print_gift      : 是否印贈品標籤
    order_total_pos : {oid: (page_idx_in_order, y)} 預先計算好的，可省去內部重掃
                      若為 None，則從 SCAN_CACHE 或重掃取得
    """
    # 統一為 list
    if isinstance(pdf_paths, str):
        pdf_paths = [pdf_paths]

    # ── 建立全域頁碼映射：global_page_idx → (pdf_path, local_page_idx) ──
    global_page_map = []   # index = global_page_idx, value = (path, local_idx)
    for path in pdf_paths:
        r = PdfReader(path)
        n = len(r.pages)
        del r
        for local_idx in range(n):
            global_page_map.append((path, local_idx))

    # ── 取得 order_total_pos ──
    if order_total_pos is None:
        # 從 SCAN_CACHE 取，沒有就掃
        combined_order_total_pos = {}
        for path in pdf_paths:
            _, otp_local = _get_or_scan(path)
            for oid, val in otp_local.items():
                if oid not in combined_order_total_pos:
                    combined_order_total_pos[oid] = val
        order_total_pos = combined_order_total_pos

    # ── 分組 ──
    groups = defaultdict(list)
    for o in orders:
        groups[o['cat']].append(o)

    TAIL_ORDER = {
        DISCOUNT_CAT: 1,
        REMARK_CAT:   2,
        SPECIAL_CAT:  3,
    }

    def sort_key(label):
        if label in TAIL_ORDER:
            return (1, TAIL_ORDER[label], 0, label)
        box_val = _GIFT_RULES.get(label, {}).get('box', '')
        if _BOX_TYPES and box_val in _BOX_TYPES:
            box_idx = _BOX_TYPES.index(box_val)
        else:
            box_idx = 9999
        return (0, 0, box_idx, label)

    sorted_labels = sorted(groups.keys(), key=sort_key)
    pad = len(str(len(sorted_labels)))

    STAMP_W  = 148
    GIFT_H   = 65
    PUR_H    = 32
    GAP      = 2
    VIP_W    = 90
    VIP_H    = 22
    AVG_W    = 95
    AVG_H    = 22
    # 統計標籤：X 範圍從 VIP 右緣到均價左緣之間置中
    # VIP 右緣 = 18 + VIP_W = 108，均價左緣 = pw_p - AVG_W - 125
    # 但 pw_p 在迴圈內才知道，這裡先定義寬度
    SUM_W    = 250   # 統計標籤寬度（points）
    SUM_H    = 40    # 統計標籤高度（auto-height 由 render 決定，這裡是最大預估）

    output_files = []

    for num, label in enumerate(sorted_labels, 1):
        ords = sort_orders_in_category(label, groups[label])

        # v4：每個分類用一個 dict 快取「已開啟的 reader」
        # key = pdf_path，value = PdfReader
        # 分類處理完統一關閉，避免重複 open/close
        open_readers: dict[str, PdfReader] = {}

        writer       = PdfWriter()
        stamp_x      = 432
        is_group_first = True
        page_counter = 0  # 分類內頁碼計數（每 50 頁 gc.collect）

        for o in ords:
            pc        = o.get('purchase_count', 0)
            pages     = o.get('page_indices', [])   # 全域頁碼
            vip_level = o.get('vip_level', '')
            avg_price = o.get('avg_price', 0)

            (target_page_idx, target_y, logi_y, subtotal_y,
             subtotal_top_y, subtotal_x1, spec_col_x0,
             subtotal_x_right, phone_bot_y) = order_total_pos.get(
                o['oid'], (0, 24, None, None, None, None, None, None, None))

            for i_in_order, global_pi in enumerate(pages):
                if global_pi >= len(global_page_map):
                    continue

                pdf_path, local_pi = global_page_map[global_pi]

                # 懶開 reader
                if pdf_path not in open_readers:
                    open_readers[pdf_path] = PdfReader(pdf_path)
                reader = open_readers[pdf_path]

                if local_pi >= len(reader.pages):
                    continue

                # ── v4 核心：mini PDF 隔離法取代 deepcopy ──
                p, mini_r, mini_buf = _isolated_page(reader, local_pi)

                ph_p = float(p.mediabox.height)
                pw_p = float(p.mediabox.width)

                is_order_first = (i_in_order == 0)
                is_total_page  = (i_in_order == target_page_idx)
                gift_y_top     = ph_p - 38.0
                gift_y         = gift_y_top - GIFT_H

                if print_gift:
                    pur_y = gift_y - GAP - PUR_H

                    if is_group_first:
                        first_o  = ords[0]
                        png_gift = render_gift_label(
                            label, first_o.get('gift', ''), first_o.get('special', False),
                            STAMP_W, GIFT_H, dpi=150
                        )
                        s = make_stamp_pdf(png_gift, pw_p, ph_p, stamp_x, gift_y, STAMP_W, GIFT_H)
                        p.merge_page(PdfReader(BytesIO(s)).pages[0])
                        if pc > 0:
                            png_pur = render_purchase_label(pc, STAMP_W, PUR_H, dpi=150)
                            s2 = make_stamp_pdf(png_pur, pw_p, ph_p, stamp_x, pur_y, STAMP_W, PUR_H)
                            p.merge_page(PdfReader(BytesIO(s2)).pages[0])
                        is_group_first = False
                    elif is_order_first and pc > 0:
                        png_pur = render_purchase_label(pc, STAMP_W, PUR_H, dpi=150)
                        s = make_stamp_pdf(png_pur, pw_p, ph_p, stamp_x, pur_y, STAMP_W, PUR_H)
                        p.merge_page(PdfReader(BytesIO(s)).pages[0])
                else:
                    if is_group_first:
                        is_group_first = False
                    if is_order_first and pc > 0:
                        pur_y_solo = ph_p - 38.0 - PUR_H
                        png_pur = render_purchase_label(pc, STAMP_W, PUR_H, dpi=150)
                        s = make_stamp_pdf(png_pur, pw_p, ph_p, stamp_x, pur_y_solo, STAMP_W, PUR_H)
                        p.merge_page(PdfReader(BytesIO(s)).pages[0])

                # VIP 標籤
                if is_total_page and vip_level:
                    vip_text = _format_vip_text(vip_level)
                    if vip_text:
                        png_vip = render_vip_label(vip_text, VIP_W, VIP_H, dpi=150)
                        vip_y = max(2, target_y - 6)
                        s = make_stamp_pdf(png_vip, pw_p, ph_p, 18, vip_y, VIP_W, VIP_H)
                        p.merge_page(PdfReader(BytesIO(s)).pages[0])

                # 均價標籤
                if is_total_page and avg_price > 0:
                    avg_text = f'均價 ${avg_price:.1f}'
                    png_avg  = render_avg_price_label(avg_text, AVG_W, AVG_H, dpi=150)
                    avg_x    = pw_p - AVG_W - 125
                    avg_y    = max(2, target_y - 6)
                    s = make_stamp_pdf(png_avg, pw_p, ph_p, avg_x, avg_y, AVG_W, AVG_H)
                    p.merge_page(PdfReader(BytesIO(s)).pages[0])

                    # 購物金標籤
                    if o.get('is_high_discount') and o.get('store_credit', 0) > 0:
                        sc_text = f'使用購物金 ${o["store_credit"]}'
                        SC_W, SC_H = AVG_W, AVG_H
                        png_sc = render_store_credit_label(sc_text, SC_W, SC_H, dpi=150)
                        sc_x = avg_x
                        sc_y = avg_y + AVG_H + 2
                        s = make_stamp_pdf(png_sc, pw_p, ph_p, sc_x, sc_y, SC_W, SC_H)
                        p.merge_page(PdfReader(BytesIO(s)).pages[0])

                # ── 商品統計標籤（印在每筆訂單第一頁）──
                if is_order_first:
                    # 順豐：X 右緣=566（小計右緣），寬328.9
                    # Y：均價框格底部往下（target_y - 6 = avg_y）
                    sx_right  = 566.0
                    s_max_w   = 328.9
                    sx_left   = sx_right - s_max_w   # 237.1
                    avg_y     = target_y - 6          # 均價框格底部 y_from_bot

                    sum_lines = _build_summary_lines(o, max_pt_w=s_max_w)
                    if sum_lines:
                        sum_pt_h = int((16 + 5) * len(sum_lines) + 12)
                        sum_y    = max(2, avg_y - sum_pt_h)
                        png_sum  = render_item_summary_label(sum_lines, s_max_w, sum_pt_h, dpi=150)
                        s = make_stamp_pdf(png_sum, pw_p, ph_p, sx_left, sum_y, s_max_w, sum_pt_h)
                        p.merge_page(PdfReader(BytesIO(s)).pages[0])


                # ── 把處理完（蓋好章）的這一頁加進輸出 writer ──
                writer.add_page(p)

                # ── v4：mini PDF 用完立刻釋放 ──
                del p, mini_r, mini_buf

                # ── v4：每 50 頁 gc.collect（Windows 記憶體回收保險）──
                page_counter += 1
                if page_counter % 50 == 0:
                    gc.collect()

        # 寫出這個分類的 PDF
        if len(writer.pages) == 0:
            # 關閉暫時開的 readers
            for r in open_readers.values():
                del r
            open_readers.clear()
            gc.collect()
            continue

        safe     = re.sub(r'[/\\ ⚠️🌀-🿿]+', '_', label).strip('_') or 'unknown'
        prefix   = str(num).zfill(pad)
        out_path = os.path.join(out_dir, f'{prefix}_{safe}.pdf')
        with open(out_path, 'wb') as f:
            writer.write(f)
        output_files.append(out_path)

        # ── 分類處理完：關閉所有 reader、釋放 writer、gc ──
        del writer
        for r in open_readers.values():
            del r
        open_readers.clear()
        gc.collect()

    clear_scan_cache()
    return output_files


def _format_vip_text(level: str) -> str:
    s = (level or '').strip().upper()
    if s == 'SVIP':
        return '【★★ SVIP】'
    if s == 'VIP':
        return '【★ VIP】'
    return ''


# ══════════════════════════════════════════════════════════════════════════════
#  蓋章相關（與 v3 完全相同）
# ══════════════════════════════════════════════════════════════════════════════
def render_gift_label(cat, gift_text, is_special, pt_w=148, pt_h=65, dpi=150):
    scale = dpi / 72.0
    W, H  = int(pt_w * scale), int(pt_h * scale)
    img   = Image.new('RGBA', (W, H), (0, 0, 0, 0))
    draw  = ImageDraw.Draw(img)
    draw.rounded_rectangle([2, 2, W-3, H-3], radius=int(6*scale),
                            fill=(255, 255, 255, 245), outline=(0, 0, 0), width=2)

    if FONT_PATH:
        f_cat  = ImageFont.truetype(FONT_PATH, int(9*scale))
        f_body = ImageFont.truetype(FONT_PATH, int(13*scale))
        f_big  = ImageFont.truetype(FONT_PATH, int(18*scale))
    else:
        f_cat = f_body = f_big = ImageFont.load_default()

    margin = int(7*scale)
    BLACK  = (0, 0, 0)

    if is_special:
        tag = {REMARK_CAT: f'【{REMARK_CAT}】', DISCOUNT_CAT: f'【{DISCOUNT_CAT}】'}.get(cat, '【特殊單】')
        draw.text((margin, int(5*scale)), tag, font=f_cat, fill=BLACK)
        draw.text((margin, int(22*scale)), '請人工確認', font=f_big, fill=BLACK)
    else:
        draw.text((margin, int(4*scale)), f'【{cat}】', font=f_cat, fill=BLACK)
        y = int(18*scale)
        for line in wrap_text(gift_text, f_body, W - margin*2)[:3]:
            draw.text((margin, y), line, font=f_body, fill=BLACK)
            y += int(17*scale)

    buf = BytesIO()
    img.save(buf, 'PNG')
    buf.seek(0)
    return buf.getvalue()


def render_purchase_label(purchase_count, pt_w=148, pt_h=32, dpi=150):
    scale = dpi / 72.0
    W, H  = int(pt_w * scale), int(pt_h * scale)
    img   = Image.new('RGBA', (W, H), (0, 0, 0, 0))
    draw  = ImageDraw.Draw(img)
    draw.rounded_rectangle([2, 2, W-3, H-3], radius=int(6*scale),
                            fill=(255, 255, 255, 245), outline=(0, 0, 0), width=2)

    if FONT_PATH:
        f_body = ImageFont.truetype(FONT_PATH, int(13*scale))
    else:
        f_body = ImageFont.load_default()

    text = f'已購買 {purchase_count} 次'
    bbox = f_body.getbbox(text)
    tw   = bbox[2] - bbox[0]
    draw.text(((W - tw) // 2, (H - (bbox[3] - bbox[1])) // 2 - int(1*scale)),
              text, font=f_body, fill=(0, 0, 0))

    buf = BytesIO()
    img.save(buf, 'PNG')
    buf.seek(0)
    return buf.getvalue()


def render_vip_label(text, pt_w=90, pt_h=22, dpi=150):
    scale = dpi / 72.0
    W, H  = int(pt_w * scale), int(pt_h * scale)
    img   = Image.new('RGBA', (W, H), (0, 0, 0, 0))
    draw  = ImageDraw.Draw(img)
    draw.rounded_rectangle([2, 2, W-3, H-3], radius=int(4*scale),
                            fill=(255, 255, 255, 245), outline=(0, 0, 0), width=2)

    if FONT_PATH:
        f = ImageFont.truetype(FONT_PATH, int(11*scale))
    else:
        f = ImageFont.load_default()

    bbox = f.getbbox(text)
    tw   = bbox[2] - bbox[0]
    draw.text(((W - tw) // 2, (H - (bbox[3] - bbox[1])) // 2 - int(1*scale)),
              text, font=f, fill=(0, 0, 0))

    buf = BytesIO()
    img.save(buf, 'PNG')
    buf.seek(0)
    return buf.getvalue()


def render_avg_price_label(text, pt_w=95, pt_h=22, dpi=150):
    scale = dpi / 72.0
    W, H  = int(pt_w * scale), int(pt_h * scale)
    img   = Image.new('RGBA', (W, H), (0, 0, 0, 0))
    draw  = ImageDraw.Draw(img)
    draw.rounded_rectangle([2, 2, W-3, H-3], radius=int(4*scale),
                            fill=(255, 255, 255, 245), outline=(0, 0, 0), width=2)

    if FONT_PATH:
        f = ImageFont.truetype(FONT_PATH, int(11*scale))
    else:
        f = ImageFont.load_default()

    bbox = f.getbbox(text)
    tw   = bbox[2] - bbox[0]
    draw.text(((W - tw) // 2, (H - (bbox[3] - bbox[1])) // 2 - int(1*scale)),
              text, font=f, fill=(0, 0, 0))

    buf = BytesIO()
    img.save(buf, 'PNG')
    buf.seek(0)
    return buf.getvalue()


def render_item_summary_label(lines: list[str], pt_w=200, pt_h=None, dpi=150):
    """
    渲染商品數量統計標籤。
    lines: 每行的文字，例如 ['C1 + C潤1', '橘2']
    pt_w : 標籤寬度（points）
    pt_h : 標籤高度（None 時自動依行數計算）
    """
    scale    = dpi / 72.0
    font_sz  = 16          # pt
    line_h   = int((font_sz + 5) * scale)
    pad_v    = int(6 * scale)
    pad_h    = int(8 * scale)

    n_lines  = max(len(lines), 1)
    if pt_h is None:
        pt_h = int((font_sz + 5) * n_lines + 12)   # auto height

    W = int(pt_w * scale)
    H = int(pt_h * scale)

    img  = Image.new('RGBA', (W, H), (0, 0, 0, 0))
    draw = ImageDraw.Draw(img)
    draw.rounded_rectangle([2, 2, W - 3, H - 3], radius=int(6 * scale),
                            fill=(255, 255, 255, 245), outline=(0, 0, 0), width=2)

    if FONT_PATH:
        f = ImageFont.truetype(FONT_PATH, int(font_sz * scale))
    else:
        f = ImageFont.load_default()

    # 文字垂直置中
    total_text_h = line_h * len(lines)
    y_cursor = max(pad_v, (H - total_text_h) // 2)
    for line in lines:
        if not line:
            y_cursor += line_h
            continue
        bbox = f.getbbox(line)
        tw   = bbox[2] - bbox[0]
        draw.text(((W - tw) // 2, y_cursor), line, font=f, fill=(0, 0, 0))
        y_cursor += line_h

    buf = BytesIO()
    img.save(buf, 'PNG')
    buf.seek(0)
    return buf.getvalue()


def _build_summary_lines(order: dict, max_pt_w: float = 328.9, dpi: int = 150) -> list[str]:
    """
    依 _SUMMARY_ORDER 統計每個子分類數量，直接用 child 名稱輸出。
    max_pt_w：每行最大寬度（points），超過自動換行。
    特別規則：_SUMMARY_ORDER 最後一組（精粹：橘/綠）強制獨立第二行，
              格式為 '+ 橘1 + 綠1'（行首固定加 +）。
    """
    child_counts: dict[str, int] = {}
    for item in order.get('items', []):
        spec = item.get('spec', '')
        qty  = item.get('qty', 1)
        info = _SPEC_MAP.get(spec)
        if not info:
            continue
        conds = info if isinstance(info, list) else [info]
        for cond in conds:
            child = cond.get('child', '')
            unit  = cond.get('qty', 1)
            if child:
                child_counts[child] = child_counts.get(child, 0) + unit * qty

    if not child_counts:
        return []

    # 精粹子分類（_SUMMARY_ORDER 最後一組）強制換行
    last_row_set = set(_SUMMARY_ORDER[-1]) if _SUMMARY_ORDER else set()

    main_items    = []
    essence_items = []
    for row in _SUMMARY_ORDER:
        for child_name in row:
            count = child_counts.get(child_name, 0)
            if count > 0:
                token = f'{child_name}{count}'
                if child_name in last_row_set:
                    essence_items.append(token)
                else:
                    main_items.append(token)

    if not main_items and not essence_items:
        return []

    # 自動換行（只對主行品項）
    try:
        from PIL import ImageFont
        scale = dpi / 72.0
        font  = ImageFont.truetype(FONT_PATH, int(16 * scale)) if FONT_PATH else None
    except Exception:
        font = None

    result: list[str] = []

    if font is None:
        if main_items:
            result.append(' + '.join(main_items))
    else:
        current: list[str] = []
        for item in main_items:
            test = ' + '.join(current + [item])
            tw   = (font.getbbox(test)[2] - font.getbbox(test)[0]) / scale + 16
            if tw > max_pt_w and current:
                result.append(' + '.join(current))
                current = [item]
            else:
                current.append(item)
        if current:
            result.append(' + '.join(current))

    # 精粹行：第一行有內容時加 + 開頭，否則直接輸出
    if essence_items:
        prefix = '+ ' if result else ''
        result.append(prefix + ' + '.join(essence_items))

    return result

def render_store_credit_label(text, pt_w=95, pt_h=22, dpi=150):
    scale = dpi / 72.0
    W, H  = int(pt_w * scale), int(pt_h * scale)
    img   = Image.new('RGBA', (W, H), (0, 0, 0, 0))
    draw  = ImageDraw.Draw(img)
    draw.rounded_rectangle([2, 2, W-3, H-3], radius=int(4*scale),
                            fill=(255, 255, 255, 245), outline=(0, 0, 0), width=2)

    if FONT_PATH:
        f = ImageFont.truetype(FONT_PATH, int(11*scale))
    else:
        f = ImageFont.load_default()

    bbox = f.getbbox(text)
    tw   = bbox[2] - bbox[0]
    draw.text(((W - tw) // 2, (H - (bbox[3] - bbox[1])) // 2 - int(1*scale)),
              text, font=f, fill=(0, 0, 0))

    buf = BytesIO()
    img.save(buf, 'PNG')
    buf.seek(0)
    return buf.getvalue()


def wrap_text(text, font, max_width):
    lines = []
    for raw in str(text).split('\n'):
        if not raw:
            lines.append('')
            continue
        while raw:
            lo, hi = 1, len(raw)
            while lo < hi:
                mid = (lo + hi + 1) // 2
                if font.getbbox(raw[:mid])[2] <= max_width:
                    lo = mid
                else:
                    hi = mid - 1
            lines.append(raw[:lo])
            raw = raw[lo:]
    return lines


def make_stamp_pdf(png_bytes, page_w, page_h, x, y, w, h):
    from reportlab.pdfgen import canvas as rl_canvas
    from reportlab.lib.utils import ImageReader
    buf = BytesIO()
    c = rl_canvas.Canvas(buf, pagesize=(page_w, page_h))
    c.drawImage(ImageReader(BytesIO(png_bytes)), x, y, width=w, height=h, mask='auto')
    c.save()
    buf.seek(0)
    return buf.getvalue()
