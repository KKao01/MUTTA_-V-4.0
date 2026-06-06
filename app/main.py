"""
分單系統 V4.1 — FastAPI 後端（合併版：7-11 + 順豐）

架構：
  - 單一入口 main.py，依前端傳入的 shipper 參數內部分流
  - shipper="seven" → 7-11 流程（V4 重繪，擷取撕線區）
  - shipper="sf"    → 順豐流程（V4 重繪，上半部留白，PDF 僅用於訂單號驗證，
                                 額外產出順豐批量下單 Excel）
  - 共用：config.json、purchase_history、core_logic、dev 後台
  - 兩套引擎並存：v4_engine（7-11）、v4_engine_sf（順豐）
"""
import os, re, json, uuid, shutil, hashlib, asyncio, tempfile, urllib.parse, csv
from pathlib import Path
from datetime import datetime
from io import BytesIO, StringIO

from fastapi import FastAPI, UploadFile, File, Form, HTTPException, Depends, Request
from fastapi.responses import StreamingResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from fastapi.middleware.cors import CORSMiddleware

BASE_DIR     = Path(__file__).parent.parent
UPLOAD_DIR   = BASE_DIR / "uploads"
CONFIG_FILE  = BASE_DIR / "config.json"
HISTORY_DIR  = BASE_DIR / "purchase_history"
SUMMARY_FILE = HISTORY_DIR / "purchase_summary.xlsx"
UPLOAD_LOG   = HISTORY_DIR / "upload_log.json"
HISTORY_NOTE = HISTORY_DIR / "note.txt"
TEMPLATE_FILE = BASE_DIR / "sf_template_clean.xlsx"   # 順豐批量下單模板
UPLOAD_DIR.mkdir(exist_ok=True)
HISTORY_DIR.mkdir(exist_ok=True)

# ══════════════════════════════════════════════════════════════════════════════
#  預設設定
# ══════════════════════════════════════════════════════════════════════════════
DEFAULT_CONFIG = {
    "dev_password_hash": hashlib.sha256(b"admin1234").hexdigest(),
    "categories": {},
    "spec_map": {},
    "classification_rules": {},
    "gift_rules": {},                # V4.1 起不再從 dev 後台 UI 維護（API 仍保留）
    "discount_excludes": [50, 120, 170],
    "discount_threshold": 157,
    "discount_min_count": 3,
    "avg_price_threshold": 450,
    "remark_excludes": ["無", "沒有", "NO", "no", "無備註"],
    "box_types": ["小空矮", "一大一小高", "一大兩小", "兩小高", "兩小矮", "2+1空"],

    # === 滿額贈整合（V4.1） ===
    "gift_builder_config": {},   # Gift Builder 完整 JSON（來自 mutta_gift_builder 匯出）
    "gift_display_map": {},      # ID → 顯示片段, 例：{"gift1": "洗髮包", "gift2": "梳"}
    "gift_fixed_suffix": "",     # 固定後綴, 例："卡"

    # 順豐固定欄位（合併自順豐 V3）
    "sf_fixed": {
        "sender_name":     "沐塔",
        "sender_phone":    "931753889",
        "sender_address":  "新北市林口區四維路329-1號",
        "sender_district": "林口區",
        "sender_city":     "新北市",
        "sender_province": "新北市",
        "sender_country":  "中國臺灣",
        "sender_postcode": "244",
        "sender_type":     "公司件",
        "sender_company":  "植沐",
        "receiver_country":  "中國臺灣",
        "receiver_province": "台灣",
        "receiver_city":     "台灣",
        "receiver_postcode": "000",
        "receiver_type":   "個人件",
        "product_name":    "保養品",
        "product_qty":     "1",
        "product_unit":    "box",
        "product_price":   "8000",
        "parcel_count":    "1",
        "currency":        "NTD",
        "express_type":    "順豐特快",
        "send_method":     "自行聯系快遞員或自寄",
    },
}


# ══════════════════════════════════════════════════════════════════════════════
#  設定管理
# ══════════════════════════════════════════════════════════════════════════════
def load_config() -> dict:
    if CONFIG_FILE.exists():
        try:
            cfg = json.loads(CONFIG_FILE.read_text(encoding="utf-8"))
            changed = False
            for key, default in DEFAULT_CONFIG.items():
                if key not in cfg:
                    cfg[key] = default
                    changed = True
            if changed:
                save_config(cfg)
            return cfg
        except Exception:
            pass
    CONFIG_FILE.write_text(
        json.dumps(DEFAULT_CONFIG, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    return DEFAULT_CONFIG.copy()


def save_config(cfg: dict):
    CONFIG_FILE.write_text(
        json.dumps(cfg, ensure_ascii=False, indent=2), encoding="utf-8"
    )


def verify_dev_password(password: str) -> bool:
    cfg = load_config()
    return hashlib.sha256(password.encode()).hexdigest() == cfg.get("dev_password_hash", "")


# ══════════════════════════════════════════════════════════════════════════════
#  歷史購買
# ══════════════════════════════════════════════════════════════════════════════
def rebuild_summary():
    """從 HISTORY_DIR 內當下所有 xlsx 檔重新計算,不 merge 舊 summary、不刪原檔。

    這樣設計可避免「同檔被重複加 N 次」的累加 bug(原本依賴 unlink 來避免重複,
    但 Windows 常因 file lock 導致 unlink 靜默失敗)。每次 rebuild 都從零重算,
    結果只跟當下資料夾內檔案有關,可重複呼叫且結果一致。
    """
    import openpyxl

    all_orders: dict = {}   # key -> set of oid (跨檔自動去重)
    for f in sorted(HISTORY_DIR.glob("*.xlsx")):
        if f.name == "purchase_summary.xlsx":
            continue
        wb = None
        try:
            wb = openpyxl.load_workbook(str(f), read_only=True)
            ws = wb["Orders"] if "Orders" in wb.sheetnames else wb.active
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            header = [str(c).strip() if c else "" for c in rows[0]]

            def get_col(col_name, row):
                try:
                    return row[header.index(col_name)]
                except (ValueError, IndexError):
                    return None

            # legacy 格式（2024/2025 舊版）沒有「對帳完成時間」欄，視同全部已對帳
            has_settled_col = "對帳完成時間" in header

            for row in rows[1:]:
                if has_settled_col:
                    settled = get_col("對帳完成時間", row)
                    if not settled or str(settled).strip() in ("", "None"):
                        continue
                oid     = str(get_col("訂單編號", row) or "").strip()
                account = str(get_col("會員帳號", row) or get_col("Email", row) or "").strip()
                name    = str(get_col("購買人姓名", row) or get_col("姓名", row) or "").strip()
                if not oid or not account or account == "None":
                    continue
                key = f"{account[:8].lower()}:{name}"
                all_orders.setdefault(key, set()).add(oid)
        except Exception:
            pass
        finally:
            try:
                if wb: wb.close()
            except Exception:
                pass

    merged = {key: len(oids) for key, oids in all_orders.items()}

    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = "summary"
    ws_out.append(["帳號key", "已購次數"])
    for key, count in sorted(merged.items()):
        ws_out.append([key, count])
    wb_out.save(str(SUMMARY_FILE))
    return len(merged)


def load_purchase_map() -> dict:
    import openpyxl
    if not SUMMARY_FILE.exists():
        return {}
    try:
        wb = openpyxl.load_workbook(str(SUMMARY_FILE), read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        result = {}
        for row in rows[1:]:
            if row[0] and row[1]:
                result[str(row[0]).strip()] = int(row[1])
        return result
    except Exception:
        return {}


# ══════════════════════════════════════════════════════════════════════════════
#  任務狀態
# ══════════════════════════════════════════════════════════════════════════════
job_store: dict[str, dict] = {}

# ══════════════════════════════════════════════════════════════════════════════
#  App
# ══════════════════════════════════════════════════════════════════════════════
app = FastAPI(title="分單系統 V4.1 (合併版)", docs_url=None, redoc_url=None)
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])

(BASE_DIR / "static").mkdir(parents=True, exist_ok=True)
app.mount("/static", StaticFiles(directory=str(BASE_DIR / "static")), name="static")
templates = Jinja2Templates(directory=str(BASE_DIR / "templates"))


@app.get("/")
async def index(request: Request):
    return templates.TemplateResponse(request, "index.html")

@app.get("/dev")
async def dev_page(request: Request):
    return templates.TemplateResponse(request, "dev.html")


# ══════════════════════════════════════════════════════════════════════════════
#  上傳：支援兩種模式
#    7-11：PDF + Excel（多檔）
#    順豐：PDF + Excel + CSV（多檔）
# ══════════════════════════════════════════════════════════════════════════════
@app.post("/api/upload")
async def upload_files(
    pdf_files:   list[UploadFile]      = File(...),
    excel_files: list[UploadFile]      = File(...),
    csv_files:   list[UploadFile]|None = File(None),
    shipper:     str                   = Form("seven"),
    show_details: bool                 = Form(True),
):
    if shipper not in ("seven", "sf"):
        raise HTTPException(400, f"未知的物流類型：{shipper}")

    # 驗證副檔名
    for f in pdf_files:
        if not f.filename.lower().endswith(".pdf"):
            raise HTTPException(400, f"只接受 PDF 檔案：{f.filename}")
    for f in excel_files:
        if not f.filename.lower().endswith((".xlsx", ".xls")):
            raise HTTPException(400, f"Excel 只接受 .xlsx / .xls 格式：{f.filename}")

    # 順豐模式：CSV 必填
    if shipper == "sf":
        if not csv_files:
            raise HTTPException(400, "順豐模式需要上傳 CSV 收件資料")
        for f in csv_files:
            if not f.filename.lower().endswith(".csv"):
                raise HTTPException(400, f"CSV 只接受 .csv 格式：{f.filename}")

    job_id = uuid.uuid4().hex

    # PDF
    pdf_list = []
    for f in pdf_files:
        raw = await f.read()
        pdf_list.append({"name": f.filename, "bytes": raw})
    pdf_name = (pdf_files[0].filename if len(pdf_files) == 1
                else f"{pdf_files[0].filename} 等 {len(pdf_files)} 個檔案")

    # Excel
    excel_list = []
    for f in excel_files:
        raw = await f.read()
        excel_list.append({"name": f.filename, "bytes": raw})
    excel_name = (excel_files[0].filename if len(excel_files) == 1
                  else f"{excel_files[0].filename} 等 {len(excel_files)} 個檔案")

    # CSV（順豐）
    csv_bytes = None
    csv_name  = None
    if shipper == "sf" and csv_files:
        if len(csv_files) == 1:
            csv_bytes = await csv_files[0].read()
            csv_name  = csv_files[0].filename
        else:
            # 多 CSV 合併
            buf = StringIO()
            writer = None
            header = None
            for f in csv_files:
                raw = await f.read()
                text = raw.decode("utf-8-sig", errors="ignore")
                reader = csv.reader(StringIO(text))
                rows = list(reader)
                if not rows: continue
                if writer is None:
                    writer = csv.writer(buf)
                    header = rows[0]
                    writer.writerow(header)
                for r in rows[1:]:
                    writer.writerow(r)
            csv_bytes = buf.getvalue().encode("utf-8-sig")
            csv_name  = f"{csv_files[0].filename} 等 {len(csv_files)} 個檔案"

    job_store[job_id] = {
        "status":      "uploaded",
        "shipper":     shipper,
        "show_details": show_details,
        "pdf_name":    pdf_name,
        "excel_name":  excel_name,
        "csv_name":    csv_name,
        "pdf_list":    pdf_list,
        "excel_list":  excel_list,
        "csv_bytes":   csv_bytes,
        "logs":        [],
    }
    return {
        "job_id":     job_id,
        "shipper":    shipper,
        "pdf_name":   pdf_name,
        "excel_name": excel_name,
        "csv_name":   csv_name,
    }


@app.post("/api/process/{job_id}")
async def process(job_id: str):
    if job_id not in job_store:
        raise HTTPException(404, "找不到此任務")
    job = job_store[job_id]
    if job["status"] == "running":
        raise HTTPException(409, "任務進行中")
    job.pop("file_bytes", None)
    job["status"]   = "running"
    job["logs"]     = []
    job["progress"] = 0
    asyncio.create_task(_run_job(job_id))
    return {"ok": True}


# ══════════════════════════════════════════════════════════════════════════════
#  分流 dispatcher
# ══════════════════════════════════════════════════════════════════════════════
async def _run_job(job_id: str):
    job = job_store[job_id]
    shipper = job.get("shipper", "seven")
    if shipper == "sf":
        await _run_job_sf(job_id)
    else:
        await _run_job_seven(job_id)


# ══════════════════════════════════════════════════════════════════════════════
#  共用：解析 CSV、注入 cl 設定、寫暫存檔等小工具
# ══════════════════════════════════════════════════════════════════════════════
def _inject_config_to_cl(cl, cfg, log_fn=None):
    cl.inject_spec_map(cfg.get("spec_map", {}))
    cl.inject_classification_rules(cfg.get("classification_rules", {}))
    cl.inject_gift_rules(cfg.get("gift_rules", {}))
    cl.inject_discount_excludes(cfg.get("discount_excludes", []))
    cl.inject_remark_excludes(cfg.get("remark_excludes", []))
    cl.inject_discount_threshold(cfg.get("discount_threshold", 157))
    cl.inject_discount_min_count(cfg.get("discount_min_count", 3))
    cl.inject_avg_price_threshold(cfg.get("avg_price_threshold", 450))
    cl.inject_box_types(cfg.get("box_types", []))
    pm = load_purchase_map()
    cl.inject_purchase_map(pm)
    if log_fn:
        if not pm:
            log_fn("⚠️ purchase_summary 為空，已購買次數將全部顯示 0", "warn")
        else:
            log_fn(f"歷史購買資料：{len(pm):,} 筆帳號記錄", "ok")
    return pm


def _parse_csv_recipients(csv_bytes: bytes) -> dict:
    """解析順豐 CSV 收件資料 → {oid: {name, phone, address}}"""
    csv_data = {}
    if not csv_bytes:
        return csv_data
    text = csv_bytes.decode("utf-8-sig", errors="ignore")
    reader = csv.DictReader(StringIO(text))
    for row in reader:
        oid = (row.get("廠商訂單編號") or "").strip()
        if not oid:
            continue
        phone = (row.get("收件人手機") or row.get("收件人電話") or "").strip()
        csv_data[oid] = {
            "name":    (row.get("收件人姓名") or "").strip(),
            "phone":   phone,
            "address": (row.get("收件人地址") or "").strip(),
        }
    return csv_data


def _fmt_phone(raw: str) -> str:
    p = re.sub(r"[^\d]", "", raw or "")
    return f"{p[:4]}-{p[4:7]}-{p[7:]}" if len(p) == 10 and p.startswith("09") else (raw or "")


def _scan_pdfs_v4(pdf_paths_list, with_recipient=True):
    """
    V4 PDF 掃描：
      - with_recipient=True（7-11）：抓 oid + order_time + recipient + phone
      - with_recipient=False（順豐）：只抓 oid 跟頁碼，用於訂單號驗證
    回傳：(pdf_info, oid_pages)
    """
    import pdfplumber
    pdf_info  = {}
    oid_pages = {}
    global_offset = 0

    for pdf_path in pdf_paths_list:
        with pdfplumber.open(pdf_path) as pdf:
            total = len(pdf.pages)
            page_meta = []

            for i, page in enumerate(pdf.pages):
                chars = page.chars or []
                text  = "".join(c.get("text", "") for c in chars)

                m_oid = re.search(r"#([A-Z0-9]{7,})(?![A-Z0-9])", text)
                oid   = m_oid.group(1) if m_oid else None
                has_content = bool(re.search(r"總計|訂單金額|訂單⾦額", text))

                meta = {"oid": oid, "has_content": has_content}
                if with_recipient:
                    m_ot  = re.search(
                        r"訂購.{0,2}期[:：]\s*(\d{4}[/-]\d{2}[/-]\d{2}\s+\d{2}:\d{2}:\d{2})", text)
                    m_rec = re.search(
                        r"收件.{1}[:：]\s*(.+?)(?:收件.{1}電話|付款|發票|\s{2,})", text)
                    m_ph  = re.search(r"收件.{1}電話[:：]\s*([\d\-]+)", text)
                    meta["order_time"] = m_ot.group(1) if m_ot else ""
                    meta["recipient"]  = m_rec.group(1).strip() if m_rec else ""
                    meta["phone"]      = _fmt_phone(m_ph.group(1)) if m_ph else ""

                page_meta.append(meta)
                try: page.flush_cache()
                except Exception: pass

            i = 0
            while i < total:
                meta = page_meta[i]
                if not meta["oid"]:
                    i += 1; continue
                oid     = meta["oid"]
                indices = [i + global_offset]
                j = i + 1
                while j < total:
                    nxt = page_meta[j]
                    if not nxt["oid"] and nxt["has_content"]:
                        indices.append(j + global_offset); j += 1
                    else:
                        break
                oid_pages[oid] = indices
                info = {
                    "page_index": i,
                    "pdf_path":   pdf_path,
                }
                if with_recipient:
                    info["order_time"] = meta.get("order_time", "")
                    info["recipient"]  = meta.get("recipient",  "")
                    info["phone"]      = meta.get("phone",      "")
                pdf_info[oid] = info
                i = j

        global_offset += total

    return pdf_info, oid_pages


def _normalize_vip(raw: str) -> str:
    s = (raw or "").strip().upper()
    return s if s in ("VIP", "SVIP") else ""


def _build_v4_qty(items: list, spec_map: dict) -> dict:
    V4_KEYS = ["C","F","C潤","F潤","抗痘沐","水光沐",
               "橘","綠","痘乳","白乳","早C","晚A"]
    qty = {k: 0 for k in V4_KEYS}
    for it in items:
        info = spec_map.get(it.get("spec", ""))
        if not info: continue
        for cond in (info if isinstance(info, list) else [info]):
            child = cond.get("child", "")
            if child in qty:
                qty[child] += cond.get("qty", 1) * it.get("qty", 1)
    return qty


def _build_item_details(items: list, spec_map: dict) -> list:
    """為每個品項展開出分類明細 cats=[[child, count], ...]，與左邊商品表同一套分類；
    供右半品項明細框的下半(B)顯示。未對應 spec_map 者 cats 為空(由引擎 fallback 原始規格)。"""
    out = []
    for it in items:
        info = spec_map.get(it.get("spec", ""))
        agg, order = {}, []
        if info:
            for cond in (info if isinstance(info, list) else [info]):
                child = cond.get("child", "")
                if not child:
                    continue
                cnt = cond.get("qty", 1) * it.get("qty", 1)
                if child in agg:
                    agg[child] += cnt
                else:
                    agg[child] = cnt; order.append(child)
        cats = [[ch, agg[ch]] for ch in order]
        out.append({"name": it.get("name", ""), "spec": it.get("spec", ""),
                    "qty": it.get("qty", 1), "cats": cats})
    return out


# ══════════════════════════════════════════════════════════════════════════════
#  7-11 流程（V4 + 撕線區擷取）
# ══════════════════════════════════════════════════════════════════════════════
async def _run_job_seven(job_id: str):
    job = job_store[job_id]

    def log(msg: str, tag: str = ""):
        ts = datetime.now().strftime("%H:%M:%S")
        job["logs"].append({"ts": ts, "msg": msg, "tag": tag})

    tmp_dir = tempfile.mkdtemp()
    try:
        import sys
        from collections import defaultdict
        from pypdf import PdfReader as _PdfReader, PdfWriter as _PdfWriter

        sys.path.insert(0, str(BASE_DIR / "app"))
        import core_logic as cl
        import v4_engine as v4
        log("引擎：超商版（v4_engine）", "ok")

        # 1. 注入設定
        cfg = load_config()
        _inject_config_to_cl(cl, cfg, log)

        # 2. 寫入暫存檔
        pdf_paths, excel_paths = [], []
        for i, item in enumerate(job["pdf_list"]):
            p = os.path.join(tmp_dir, f"input_{i:02d}.pdf")
            with open(p, "wb") as f: f.write(item["bytes"])
            pdf_paths.append(p)
        for i, item in enumerate(job["excel_list"]):
            p = os.path.join(tmp_dir, f"orders_{i:02d}.xlsx")
            with open(p, "wb") as f: f.write(item["bytes"])
            excel_paths.append(p)

        job["progress"] = 5
        log(f"載入 PDF：{job['pdf_name']}")
        log(f"載入 Excel：{job['excel_name']}")

        loop = asyncio.get_event_loop()

        # 3. 解析 Excel
        orders = await loop.run_in_executor(None, cl.parse_excels, excel_paths)
        job["progress"] = 20
        log(f"Excel 解析完成：{len(orders)} 筆訂單", "ok")

        # 4. 掃描 PDF（含收件人/電話/訂購日期）
        log("掃描 PDF 頁碼與收件資訊...")
        pdf_info, oid_pages = await loop.run_in_executor(
            None, _scan_pdfs_v4, pdf_paths, True
        )
        job["progress"] = 40
        log(f"PDF 識別 {len(oid_pages)} 筆訂單", "ok")

        # 5. 比對訂單號
        excel_oids       = {o["oid"] for o in orders}
        pdf_oids         = set(oid_pages.keys())
        missing_in_pdf   = excel_oids - pdf_oids
        missing_in_excel = pdf_oids   - excel_oids

        if missing_in_pdf:
            log(f"⚠️ Excel 有但 PDF 找不到：{', '.join(sorted(missing_in_pdf))}", "warn")
        if missing_in_excel:
            log(f"⚠️ PDF 有但 Excel 找不到：{', '.join(sorted(missing_in_excel))}", "warn")
        if missing_in_pdf or missing_in_excel:
            log("❌ 訂單號碼不完全吻合，請確認 PDF 與 Excel 是否匹配", "err")
            job["status"] = "error"
            job["error"]  = "PDF 與 Excel 訂單號碼不完全吻合"
            return

        log("✓ 訂單號碼完全吻合", "ok")
        job["progress"] = 50

        # ── 5.5 解析 #MG: 贈品標籤（V4.1 新增） ──
        import gift_parser as gp
        gb_cfg        = cfg.get("gift_builder_config", {})
        display_map   = cfg.get("gift_display_map", {})
        fixed_suffix  = cfg.get("gift_fixed_suffix", "")
        fallback_map  = {g["id"]: g.get("name", g["id"])
                         for g in gb_cfg.get("gifts", [])}

        gift_hit = 0
        for o in orders:
            clean_remark, gift_display, has_gift = gp.process_order_remark(
                o.get("remark", ""),
                display_map,
                fixed_suffix,
                fallback_map,
            )
            o["remark"]       = clean_remark   # 覆寫成乾淨備註（避免被分到備註訂單）
            o["gift_display"] = gift_display   # 給 v4_engine 畫贈品欄用
            if has_gift:
                gift_hit += 1
        log(f"贈品解析完成：{gift_hit} 筆訂單含 #MG: 標籤", "ok" if gift_hit else "")

        # 6. 分類
        log("進行分類判斷...")
        for o in orders:
            cat, is_special, is_tail = cl.classify_order(o)
            o["cat"]     = cat
            o["special"] = is_special
            o["is_tail"] = is_tail
            sex          = cl.gender(o.get("name", ""))
            o["gift"]    = cl.get_gift(cat, sex) if not is_special else "請人工確認"

        cl.reassign_special_categories(orders)
        for o in orders:
            if o.get("cat") in (cl.REMARK_CAT, cl.DISCOUNT_CAT, cl.SPECIAL_CAT):
                o["special"] = True
                o["gift"]    = "請人工確認"

        pc_hit = sum(1 for o in orders if o.get("purchase_count", 0) > 0)
        log(f"分類完成（{pc_hit} 筆訂單有購買記錄）", "ok")
        job["progress"] = 60

        # 7. 排序、組合 V4 dict、輸出 PDF
        sorted_labels, v4_orders = _build_v4_orders(orders, cfg, pdf_info, pdf_paths)
        for _o in v4_orders:
            _o["show_details"] = job.get("show_details", True)
        job["progress"] = 75

        log("輸出 V4 分類 PDF...")
        v4.setup_font()
        output_files = await loop.run_in_executor(
            None, _generate_v4_pdfs, v4, v4_orders, sorted_labels, tmp_dir
        )
        job["progress"] = 95

        # 8. 收集結果
        _collect_results(job, job_id, output_files, orders, log)

        job["status"]   = "done"
        job["progress"] = 100
        log(f"完成！共 {len(output_files)} 個分類 PDF，{len(v4_orders)} 頁", "ok")

    except Exception as e:
        job["status"] = "error"
        job["error"]  = str(e)
        log(f"錯誤：{e}", "err")
        import traceback; traceback.print_exc()
    finally:
        shutil.rmtree(tmp_dir, ignore_errors=True)


# ══════════════════════════════════════════════════════════════════════════════
#  順豐流程（V4 + 上半部留白 + CSV + 順豐批量下單 Excel）
# ══════════════════════════════════════════════════════════════════════════════
async def _run_job_sf(job_id: str):
    job = job_store[job_id]

    def log(msg: str, tag: str = ""):
        ts = datetime.now().strftime("%H:%M:%S")
        job["logs"].append({"ts": ts, "msg": msg, "tag": tag})

    tmp_dir = tempfile.mkdtemp()
    try:
        import sys
        from collections import defaultdict
        from pypdf import PdfReader as _PdfReader, PdfWriter as _PdfWriter

        sys.path.insert(0, str(BASE_DIR / "app"))
        import core_logic as cl
        import v4_engine_sf as v4
        log("引擎：順豐版（v4_engine_sf）", "ok")

        # 1. 注入設定
        cfg = load_config()
        _inject_config_to_cl(cl, cfg, log)

        # 2. 寫入暫存檔
        pdf_paths, excel_paths = [], []
        for i, item in enumerate(job["pdf_list"]):
            p = os.path.join(tmp_dir, f"input_{i:02d}.pdf")
            with open(p, "wb") as f: f.write(item["bytes"])
            pdf_paths.append(p)
        for i, item in enumerate(job["excel_list"]):
            p = os.path.join(tmp_dir, f"orders_{i:02d}.xlsx")
            with open(p, "wb") as f: f.write(item["bytes"])
            excel_paths.append(p)

        job["progress"] = 5
        log(f"載入 PDF：{job['pdf_name']}")
        log(f"載入 Excel：{job['excel_name']}")
        log(f"載入 CSV：{job['csv_name']}")

        loop = asyncio.get_event_loop()

        # 3. 解析 Excel
        orders = await loop.run_in_executor(None, cl.parse_excels, excel_paths)
        job["progress"] = 18
        log(f"Excel 解析完成：{len(orders)} 筆訂單", "ok")

        # 4. 解析 CSV（收件資料）
        csv_data = _parse_csv_recipients(job.get("csv_bytes"))
        job["csv_data"] = csv_data
        log(f"CSV 解析完成：{len(csv_data)} 筆收件資料", "ok")
        job["progress"] = 25

        # 5. 掃描 PDF（順豐版只抓 oid 用於驗證，不抓收件資訊）
        log("掃描 PDF 頁碼（僅用於訂單號驗證）...")
        pdf_info, oid_pages = await loop.run_in_executor(
            None, _scan_pdfs_v4, pdf_paths, False
        )
        job["progress"] = 40
        log(f"PDF 識別 {len(oid_pages)} 筆訂單", "ok")

        # 6. 三方比對：Excel ⇄ PDF ⇄ CSV
        excel_oids = {o["oid"] for o in orders}
        pdf_oids   = set(oid_pages.keys())
        csv_oids   = set(csv_data.keys())

        missing_in_pdf   = excel_oids - pdf_oids
        missing_in_excel = pdf_oids   - excel_oids
        missing_in_csv   = excel_oids - csv_oids

        if missing_in_pdf:
            log(f"⚠️ Excel 有但 PDF 找不到：{', '.join(sorted(missing_in_pdf))}", "warn")
        if missing_in_excel:
            log(f"⚠️ PDF 有但 Excel 找不到：{', '.join(sorted(missing_in_excel))}", "warn")
        if missing_in_csv:
            log(f"⚠️ Excel 有但 CSV 找不到：{', '.join(sorted(missing_in_csv))}", "warn")

        if missing_in_pdf or missing_in_excel:
            log("❌ PDF 與 Excel 訂單號碼不完全吻合", "err")
            job["status"] = "error"
            job["error"]  = "PDF 與 Excel 訂單號碼不完全吻合"
            return

        log("✓ 訂單號碼驗證通過", "ok")
        job["progress"] = 50

        # ── 6.5 解析 #MG: 贈品標籤（V4.1 新增） ──
        import gift_parser as gp
        gb_cfg        = cfg.get("gift_builder_config", {})
        display_map   = cfg.get("gift_display_map", {})
        fixed_suffix  = cfg.get("gift_fixed_suffix", "")
        fallback_map  = {g["id"]: g.get("name", g["id"])
                         for g in gb_cfg.get("gifts", [])}

        gift_hit = 0
        for o in orders:
            clean_remark, gift_display, has_gift = gp.process_order_remark(
                o.get("remark", ""),
                display_map,
                fixed_suffix,
                fallback_map,
            )
            o["remark"]       = clean_remark   # 覆寫成乾淨備註（避免被分到備註訂單）
            o["gift_display"] = gift_display   # 給 v4_engine 畫贈品欄用
            if has_gift:
                gift_hit += 1
        log(f"贈品解析完成：{gift_hit} 筆訂單含 #MG: 標籤", "ok" if gift_hit else "")

        # 7. 分類
        log("進行分類判斷...")
        for o in orders:
            cat, is_special, is_tail = cl.classify_order(o)
            o["cat"]     = cat
            o["special"] = is_special
            o["is_tail"] = is_tail
            sex          = cl.gender(o.get("name", ""))
            o["gift"]    = cl.get_gift(cat, sex) if not is_special else "請人工確認"

        cl.reassign_special_categories(orders)
        for o in orders:
            if o.get("cat") in (cl.REMARK_CAT, cl.DISCOUNT_CAT, cl.SPECIAL_CAT):
                o["special"] = True
                o["gift"]    = "請人工確認"

        pc_hit = sum(1 for o in orders if o.get("purchase_count", 0) > 0)
        log(f"分類完成（{pc_hit} 筆訂單有購買記錄）", "ok")
        job["progress"] = 60

        # 8. 排序、組合 V4 dict（收件人/電話以 CSV 為主）
        sorted_labels, v4_orders = _build_v4_orders(
            orders, cfg, pdf_info, pdf_paths, csv_data=csv_data
        )
        for _o in v4_orders:
            _o["show_details"] = job.get("show_details", True)
        job["progress"] = 75

        # 9. 輸出 PDF（用 v4_engine_sf）
        log("輸出 V4 分類 PDF（順豐版）...")
        v4.setup_font()
        output_files = await loop.run_in_executor(
            None, _generate_v4_pdfs, v4, v4_orders, sorted_labels, tmp_dir
        )
        job["progress"] = 88

        # 10. 收集結果
        _collect_results(job, job_id, output_files, orders, log)

        # 11. 產出順豐批量下單 Excel
        log("產生順豐批量下單 Excel...")
        # ordered_for_excel 跟 PDF 順序一致
        ordered_for_excel = []
        groups_by_cat = {}
        for o in orders:
            groups_by_cat.setdefault(o["cat"], []).append(o)
        for label in sorted_labels:
            ords_in_cat = cl.sort_orders_in_category(label, groups_by_cat[label])
            ordered_for_excel.extend(ords_in_cat)

        sf_fixed = cfg.get("sf_fixed", {})
        try:
            sf_excel_bytes = _build_sf_excel(ordered_for_excel, csv_data, sf_fixed)
            date_str = datetime.now().strftime("%Y%m%d")
            sf_excel_name = f"順豐批量下單_{date_str}.xlsx"
            job["file_bytes"][sf_excel_name] = sf_excel_bytes
            log(f"順豐 Excel：{len(ordered_for_excel)} 筆", "ok")
        except HTTPException as he:
            log(f"⚠️ 順豐 Excel 跳過：{he.detail}", "warn")
        except Exception as e:
            log(f"⚠️ 順豐 Excel 產生失敗：{e}", "warn")

        job["status"]   = "done"
        job["progress"] = 100
        log(f"完成！{len(output_files)} 個分類 PDF + 順豐批量下單 Excel", "ok")

    except Exception as e:
        job["status"] = "error"
        job["error"]  = str(e)
        log(f"錯誤：{e}", "err")
        import traceback; traceback.print_exc()
    finally:
        shutil.rmtree(tmp_dir, ignore_errors=True)


# ══════════════════════════════════════════════════════════════════════════════
#  共用：build v4_orders、generate_v4_pdfs、collect_results
# ══════════════════════════════════════════════════════════════════════════════
def _build_v4_orders(orders, cfg, pdf_info, pdf_paths, csv_data=None):
    """
    將 V3 訂單 dict 轉成 V4 引擎需要的格式。
    csv_data: 順豐 CSV 收件資料 → 收件人/電話以 CSV 為主，PDF 為備援
    """
    from collections import defaultdict
    groups = defaultdict(list)
    for o in orders:
        groups[o["cat"]].append(o)

    # 引入 cl 取常數
    import sys
    sys.path.insert(0, str(BASE_DIR / "app"))
    import core_logic as cl

    TAIL_ORDER = {cl.DISCOUNT_CAT: 1, cl.REMARK_CAT: 2, cl.SPECIAL_CAT: 3}
    gift_rules = cfg.get("gift_rules", {})
    box_types  = cfg.get("box_types", [])

    def _sort_key(label):
        if label in TAIL_ORDER:
            return (1, TAIL_ORDER[label], 0, label)
        box_val = gift_rules.get(label, {}).get("box", "")
        box_idx = box_types.index(box_val) if (box_types and box_val in box_types) else 9999
        return (0, 0, box_idx, label)

    sorted_labels = sorted(groups.keys(), key=_sort_key)
    v3_ordered = []
    for label in sorted_labels:
        v3_ordered.extend(cl.sort_orders_in_category(label, groups[label]))

    print_time = datetime.now().strftime("%Y/%m/%d %H:%M:%S")
    threshold  = cfg.get("discount_threshold", 157)
    excludes   = set(cfg.get("discount_excludes", []))

    v4_orders = []
    for v3o in v3_ordered:
        sc     = v3o.get("store_credit", 0)
        wallet = sc if (sc > threshold and sc not in excludes) else None
        avg    = v3o.get("avg_price", 0)
        info   = pdf_info.get(v3o["oid"], {})

        # 訂購人 / 收貨人 / 電話 / 物流 / 訂購日期 全部從 Excel 來
        # CSV（順豐）和 PDF 抓的當 fallback
        buyer = v3o.get("buyer") or v3o.get("name", "")

        if csv_data is not None:
            csv_row   = csv_data.get(v3o["oid"], {})
            recipient = csv_row.get("name")  or v3o.get("recipient") or info.get("recipient") or v3o.get("name", "")
            phone     = csv_row.get("phone") or v3o.get("phone")     or info.get("phone", "")
        else:
            recipient = v3o.get("recipient") or info.get("recipient") or v3o.get("name", "")
            phone     = v3o.get("phone")     or info.get("phone", "")
        phone = _fmt_phone(phone) if phone else ""

        shipping_method = v3o.get("shipping_method", "")
        order_time = v3o.get("order_time") or info.get("order_time", "")

        v4o = {
            "order_id":         v3o["oid"],
            "print_time":       print_time,
            "order_time":       order_time,
            "invoice_time":     None,
            "buyer":            buyer,
            "recipient":        recipient,
            "phone":            phone,
            "shipping_method":  shipping_method,
            "remark":           v3o.get("remark", ""),
            "vip_level":        _normalize_vip(v3o.get("vip_level", "")),
            "purchase_count":   v3o.get("purchase_count", 0),
            "wallet":           wallet,
            "avg_price":        round(avg, 1) if avg and avg > 0 else None,
            "subtotal":         v3o.get("order_total", 0),
            "product_discount": -abs(v3o.get("discount", 0)) if v3o.get("discount") else 0,
            "order_discount":   0,
            "shipping":         v3o.get("shipping", 0),
            "adjustment":       0,
            "total":            v3o.get("order_total", 0),
            "qty":              _build_v4_qty(v3o.get("items", []), cfg.get("spec_map", {})),
            "items":            _build_item_details(v3o.get("items", []), cfg.get("spec_map", {})),
            "_cat":             v3o.get("cat", "特殊單"),
            "_page_index":      info.get("page_index", 0),
            "_pdf_path":        info.get("pdf_path", pdf_paths[0] if pdf_paths else ""),
            "gift_display":     v3o.get("gift_display", ""),
        }
        v4_orders.append(v4o)

    return sorted_labels, v4_orders


def _generate_v4_pdfs(v4, v4_orders, sorted_labels, tmp_dir):
    """執行 V4 引擎產出每個分類的 PDF"""
    from pypdf import PdfReader as _PdfReader, PdfWriter as _PdfWriter
    pad = len(str(len(sorted_labels)))
    output = []
    global_i = 0
    for num, label in enumerate(sorted_labels, 1):
        cat_v4 = [o for o in v4_orders if o.get("_cat") == label]
        if not cat_v4: continue

        writer     = _PdfWriter()
        safe_label = label.replace("/", "_").replace("\\", "_")
        out_path   = os.path.join(tmp_dir, f"{num:0{pad}d}_{safe_label}.pdf")

        for order in cat_v4:
            global_i += 1
            page_idx  = order.pop("_page_index", 0)
            pdf_path  = order.pop("_pdf_path", "")
            _cat      = order.pop("_cat", "")
            tmp_path  = os.path.join(tmp_dir, f"v4_{global_i:04d}.pdf")
            v4.generate_page(
                pdf_path    = pdf_path,
                page_index  = page_idx,
                order       = order,
                output_path = tmp_path,
            )
            r = _PdfReader(tmp_path)
            writer.add_page(r.pages[0])
            del r

        with open(out_path, "wb") as fh:
            writer.write(fh)
        output.append(out_path)
    return output


def _collect_results(job, job_id, output_files, orders, log):
    """收集結果到 job dict"""
    from collections import defaultdict
    groups_by_cat = defaultdict(list)
    for o in orders:
        groups_by_cat[o["cat"]].append(o)

    file_bytes = {}
    results    = []
    for path in output_files:
        fname = os.path.basename(path)
        file_bytes[fname] = Path(path).read_bytes()
        label = re.sub(r"^\d+_", "", os.path.splitext(fname)[0])
        cat_ords = groups_by_cat.get(label, [])
        is_sp = cat_ords[0].get("special", False) if cat_ords else False
        results.append({
            "cat":      label,
            "filename": fname,
            "count":    len(cat_ords),
            "pages":    len(cat_ords),
            "gift":     cat_ords[0].get("gift", "") if cat_ords else "",
            "special":  is_sp,
            "url":      f"/api/download/{job_id}/{fname}",
        })
        log(f"[{label}]  {len(cat_ords)} 筆", "warn" if is_sp else "ok")

    job["file_bytes"] = file_bytes
    job["results"]    = results


# ══════════════════════════════════════════════════════════════════════════════
#  順豐批量下單 Excel（從順豐 V3 移植）
# ══════════════════════════════════════════════════════════════════════════════
def _build_sf_excel(ordered: list, csv_data: dict, sf_fixed: dict) -> bytes:
    from openpyxl import load_workbook
    if not TEMPLATE_FILE.exists():
        raise HTTPException(status_code=500, detail="找不到順豐模板 sf_template_clean.xlsx")

    wb = load_workbook(str(TEMPLATE_FILE))
    ws = wb["information"]

    col_map = {}
    for c in range(1, ws.max_column + 1):
        val = ws.cell(1, c).value
        if val:
            col_map[val] = c

    def set_col(col_name, value, row_idx):
        c = col_map.get(col_name)
        if c:
            ws.cell(row_idx, c).value = value

    for row_idx, order in enumerate(ordered, start=3):
        oid     = order.get("oid", "")
        csv_row = csv_data.get(oid, {})

        set_col("*客戶訂單號",      oid,                              row_idx)
        set_col("*收件方姓名",      csv_row.get("name", ""),          row_idx)
        set_col("收件方手機號",     csv_row.get("phone", ""),         row_idx)
        set_col("*收件方詳細地址",  csv_row.get("address", ""),       row_idx)
        set_col("*寄件方姓名",      sf_fixed.get("sender_name", ""),     row_idx)
        set_col("寄件方手機號",     sf_fixed.get("sender_phone", ""),    row_idx)
        set_col("*寄件方詳細地址",  sf_fixed.get("sender_address", ""),  row_idx)
        set_col("寄件方縣/區",      sf_fixed.get("sender_district", ""), row_idx)
        set_col("*寄件方城市",      sf_fixed.get("sender_city", ""),     row_idx)
        set_col("*寄件方州/省",     sf_fixed.get("sender_province", ""), row_idx)
        set_col("*寄件方國家/地區", sf_fixed.get("sender_country", ""),  row_idx)
        set_col("*寄件方郵編",      sf_fixed.get("sender_postcode", ""), row_idx)
        set_col("寄件類型",         sf_fixed.get("sender_type", ""),     row_idx)
        set_col("寄件方公司",       sf_fixed.get("sender_company", ""),  row_idx)
        set_col("*收件方國家/地區", sf_fixed.get("receiver_country", ""),  row_idx)
        set_col("*收件方州/省",     sf_fixed.get("receiver_province", ""), row_idx)
        set_col("*收件方城市",      sf_fixed.get("receiver_city", ""),     row_idx)
        set_col("*收件方郵編",      sf_fixed.get("receiver_postcode", ""), row_idx)
        set_col("收件類型",         sf_fixed.get("receiver_type", ""),     row_idx)
        set_col("*商品名稱",        sf_fixed.get("product_name", ""),  row_idx)
        set_col("*商品數量",        sf_fixed.get("product_qty", ""),   row_idx)
        set_col("*單位",            sf_fixed.get("product_unit", ""),  row_idx)
        set_col("*商品單價",        sf_fixed.get("product_price", ""), row_idx)
        set_col("*包裹總件數",      sf_fixed.get("parcel_count", ""),  row_idx)
        set_col("*商品貨幣",        sf_fixed.get("currency", ""),      row_idx)
        set_col("*快件類型",        sf_fixed.get("express_type", ""),  row_idx)
        set_col("*寄件方式",        sf_fixed.get("send_method", ""),   row_idx)
        set_col("*付款方式",        sf_fixed.get("payment_method", ""),  row_idx)
        set_col("月結卡號",          sf_fixed.get("monthly_card_no", ""), row_idx)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ══════════════════════════════════════════════════════════════════════════════
#  查詢進度 / 下載
# ══════════════════════════════════════════════════════════════════════════════
@app.get("/api/status/{job_id}")
async def job_status(job_id: str):
    if job_id not in job_store:
        raise HTTPException(404, "找不到此任務")
    return {k: v for k, v in job_store[job_id].items()
            if k not in ("file_bytes", "pdf_list", "excel_list", "csv_bytes", "csv_data")}


def _download_response(job_id: str, filename: str):
    job = job_store.get(job_id)
    if not job or filename not in job.get("file_bytes", {}):
        raise HTTPException(404, "檔案不存在")
    data    = job["file_bytes"][filename]
    encoded = urllib.parse.quote(filename, safe="")
    media   = ("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
               if filename.endswith(".xlsx") else "application/pdf")
    return StreamingResponse(
        BytesIO(data), media_type=media,
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded}"}
    )

@app.get("/api/download/{job_id}")
async def download_file_query(job_id: str, filename: str):
    return _download_response(job_id, filename)

@app.get("/api/download/{job_id}/{filename}")
async def download_file(job_id: str, filename: str):
    return _download_response(job_id, filename)

@app.get("/api/download-all/{job_id}")
async def download_all(job_id: str):
    import zipfile
    job = job_store.get(job_id)
    if not job or not job.get("file_bytes"):
        raise HTTPException(404, "找不到任務或檔案")
    buf = BytesIO()
    with zipfile.ZipFile(buf, "w") as zf:
        for fname, data in job["file_bytes"].items():
            zf.writestr(fname, data)
    buf.seek(0)
    prefix = "順豐" if job.get("shipper") == "sf" else "超商"
    zipname = urllib.parse.quote(f'{prefix}_{datetime.now().strftime("%Y%m%d")}.zip')
    return StreamingResponse(
        buf, media_type="application/zip",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{zipname}"}
    )

@app.post("/api/cleanup/{job_id}")
async def cleanup_job(job_id: str):
    job_store.pop(job_id, None)
    return {"ok": True}


# ══════════════════════════════════════════════════════════════════════════════
#  開發者 API
# ══════════════════════════════════════════════════════════════════════════════
def require_dev_token(request: Request):
    token = request.headers.get("X-Dev-Token", "")
    if not token or not verify_dev_password(token):
        raise HTTPException(401, "未授權")

@app.post("/api/dev/login")
async def dev_login(password: str = Form(...)):
    if not verify_dev_password(password):
        raise HTTPException(401, "密碼錯誤")
    return {"ok": True, "token": password}

@app.get("/api/dev/config")
async def get_config(request: Request, _=Depends(require_dev_token)):
    return load_config()

@app.get("/api/dev/categories")
async def get_categories(request: Request, _=Depends(require_dev_token)):
    return load_config().get("categories", {})

@app.post("/api/dev/categories")
async def save_categories(request: Request, _=Depends(require_dev_token)):
    body = await request.json()
    cfg = load_config(); cfg["categories"] = body; save_config(cfg)
    return {"ok": True}

@app.get("/api/dev/spec-map")
async def get_spec_map(request: Request, _=Depends(require_dev_token)):
    return load_config().get("spec_map", {})

@app.post("/api/dev/spec-map")
async def save_spec_map(request: Request, _=Depends(require_dev_token)):
    body = await request.json()
    cfg = load_config(); cfg["spec_map"] = body; save_config(cfg)
    return {"ok": True}

@app.post("/api/dev/parse-product-excel")
async def parse_product_excel(
    request: Request, file: UploadFile = File(...), _=Depends(require_dev_token)
):
    if not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(400, "只接受 .xlsx / .xls 檔案")
    try:
        import openpyxl
        content = await file.read()
        wb = openpyxl.load_workbook(BytesIO(content), read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows: raise HTTPException(400, "檔案為空")
        header = [str(c).strip() if c else "" for c in rows[0]]
        def get_col(name, row):
            try:
                idx = header.index(name); v = row[idx]
                return str(v).strip() if v and str(v).strip() not in ("None","nan","NaN","") else ""
            except (ValueError, IndexError): return ""
        cfg = load_config(); existing = cfg.get("spec_map", {})
        results = []; seen = set()
        for row in rows[1:]:
            product_name = get_col("商品名稱", row)
            if not product_name: continue
            p=get_col("品項一",row); q=get_col("品項二",row)
            r=get_col("組合擇一",row); s=get_col("品項",row); t=get_col("商品",row)
            if p and q: spec_key = f"{p}/{q}"
            elif r: spec_key = r
            elif s: spec_key = s
            elif t: spec_key = t
            else: continue
            if spec_key in seen: continue
            seen.add(spec_key)
            ex_raw = existing.get(spec_key)
            ex = (ex_raw[0] if isinstance(ex_raw,list) and ex_raw
                  else ex_raw if isinstance(ex_raw,dict) else {})
            results.append({"spec":spec_key,"product_name":product_name,
                            "parent":ex.get("parent",""),"child":ex.get("child",""),
                            "qty":ex.get("qty",1),"is_new":spec_key not in existing})
        return {"rows": results, "total": len(results)}
    except HTTPException: raise
    except Exception as e: raise HTTPException(500, f"解析失敗：{e}")

@app.get("/api/dev/classification-rules")
async def get_classification_rules(request: Request, _=Depends(require_dev_token)):
    return load_config().get("classification_rules", {})

@app.post("/api/dev/classification-rules")
async def save_classification_rules(request: Request, _=Depends(require_dev_token)):
    body = await request.json()
    cfg = load_config(); cfg["classification_rules"] = body; save_config(cfg)
    return {"ok": True}

@app.get("/api/dev/gift-rules")
async def get_gift_rules(request: Request, _=Depends(require_dev_token)):
    return load_config().get("gift_rules", {})

@app.post("/api/dev/gift-rules")
async def save_gift_rules(request: Request, _=Depends(require_dev_token)):
    body = await request.json()
    cfg = load_config(); cfg["gift_rules"] = body; save_config(cfg)
    return {"ok": True}

# === 滿額贈整合（V4.1）===
@app.get("/api/dev/gift-builder-config")
async def get_gift_builder_config(request: Request, _=Depends(require_dev_token)):
    return load_config().get("gift_builder_config", {})

@app.post("/api/dev/gift-builder-config")
async def save_gift_builder_config(request: Request, _=Depends(require_dev_token)):
    body = await request.json()
    cfg = load_config()
    cfg["gift_builder_config"] = body
    save_config(cfg)
    return {"ok": True}

@app.get("/api/dev/gift-display-map")
async def get_gift_display_map(request: Request, _=Depends(require_dev_token)):
    cfg = load_config()
    return {
        "display_map":   cfg.get("gift_display_map", {}),
        "fixed_suffix":  cfg.get("gift_fixed_suffix", ""),
    }

@app.post("/api/dev/gift-display-map")
async def save_gift_display_map(request: Request, _=Depends(require_dev_token)):
    body = await request.json()
    cfg = load_config()
    cfg["gift_display_map"]  = body.get("display_map", {})
    cfg["gift_fixed_suffix"] = body.get("fixed_suffix", "")
    save_config(cfg)
    return {"ok": True}

@app.post("/api/dev/discount-excludes")
async def update_discount_excludes(request: Request, _=Depends(require_dev_token)):
    body = await request.json()
    cfg = load_config(); cfg["discount_excludes"] = body.get("values",[]); save_config(cfg)
    return {"ok": True}

@app.get("/api/dev/discount-config")
async def get_discount_config(request: Request, _=Depends(require_dev_token)):
    cfg = load_config()
    return {"discount_threshold": cfg.get("discount_threshold",157),
            "discount_min_count": cfg.get("discount_min_count",3),
            "avg_price_threshold": cfg.get("avg_price_threshold",450)}

@app.post("/api/dev/discount-config")
async def save_discount_config(request: Request, _=Depends(require_dev_token)):
    body = await request.json(); cfg = load_config()
    try:
        cfg["discount_threshold"]  = int(body.get("discount_threshold",  157))
        cfg["discount_min_count"]  = int(body.get("discount_min_count",  3))
        cfg["avg_price_threshold"] = int(body.get("avg_price_threshold", 450))
    except (ValueError, TypeError): raise HTTPException(400, "數值格式錯誤")
    save_config(cfg); return {"ok": True}

@app.post("/api/dev/remark-excludes")
async def update_remark_excludes(request: Request, _=Depends(require_dev_token)):
    body = await request.json()
    cfg = load_config(); cfg["remark_excludes"] = body.get("values",[]); save_config(cfg)
    return {"ok": True}

@app.get("/api/dev/box-types")
async def get_box_types(request: Request, _=Depends(require_dev_token)):
    return load_config().get("box_types", [])

@app.post("/api/dev/box-types")
async def save_box_types(request: Request, _=Depends(require_dev_token)):
    body = await request.json()
    cfg = load_config(); cfg["box_types"] = body; save_config(cfg)
    return {"ok": True}

# ── 順豐固定欄位 API（從順豐 V3 移植）──
@app.get("/api/dev/sf-fixed")
async def get_sf_fixed(request: Request, _=Depends(require_dev_token)):
    return load_config().get("sf_fixed", DEFAULT_CONFIG["sf_fixed"])

@app.post("/api/dev/sf-fixed")
async def save_sf_fixed(request: Request, _=Depends(require_dev_token)):
    body = await request.json()
    cfg = load_config(); cfg["sf_fixed"] = body; save_config(cfg)
    return {"ok": True}

@app.post("/api/dev/change-password")
async def change_password(request: Request, _=Depends(require_dev_token)):
    body = await request.json()
    cur = body.get("current",""); new = body.get("new","")
    if not verify_dev_password(cur): raise HTTPException(401, "目前密碼錯誤")
    if len(new) < 6: raise HTTPException(400, "新密碼至少 6 碼")
    cfg = load_config()
    cfg["dev_password_hash"] = hashlib.sha256(new.encode()).hexdigest()
    save_config(cfg); return {"ok": True}

@app.get("/api/dev/export")
async def export_config(request: Request, _=Depends(require_dev_token)):
    cfg = load_config()
    export_data = {
        "exported_at":          datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "categories":           cfg.get("categories",           {}),
        "spec_map":             cfg.get("spec_map",             {}),
        "classification_rules": cfg.get("classification_rules", {}),
        "gift_rules":           cfg.get("gift_rules",           {}),
        "discount_excludes":    cfg.get("discount_excludes",    []),
        "discount_threshold":   cfg.get("discount_threshold",   157),
        "discount_min_count":   cfg.get("discount_min_count",   3),
        "avg_price_threshold":  cfg.get("avg_price_threshold",  450),
        "remark_excludes":      cfg.get("remark_excludes",      []),
        "box_types":            cfg.get("box_types",            []),
        "sf_fixed":             cfg.get("sf_fixed",             {}),
    }
    content  = json.dumps(export_data, ensure_ascii=False, indent=2)
    filename = f"v4_config_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
    return StreamingResponse(BytesIO(content.encode("utf-8")),
        media_type="application/json",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"})

@app.post("/api/dev/import")
async def import_config(request: Request, file: UploadFile = File(...), _=Depends(require_dev_token)):
    if not file.filename.lower().endswith(".json"):
        raise HTTPException(400, "只接受 .json 檔案")
    try:
        raw = await file.read(); data = json.loads(raw.decode("utf-8"))
    except Exception: raise HTTPException(400, "JSON 格式錯誤")
    cfg = load_config()
    for key in ("categories","spec_map","classification_rules","gift_rules",
                "discount_excludes","discount_threshold","discount_min_count",
                "avg_price_threshold","remark_excludes","box_types","sf_fixed"):
        if key in data: cfg[key] = data[key]
    save_config(cfg)
    return {"ok": True, "imported_at": data.get("exported_at", "")}

# ── 歷史購買 API ──────────────────────────────────────────────────────────────
@app.get("/api/dev/history/files")
async def history_list(request: Request, _=Depends(require_dev_token)):
    pm = load_purchase_map()
    log_data = []
    if UPLOAD_LOG.exists():
        try: log_data = json.loads(UPLOAD_LOG.read_text(encoding="utf-8"))
        except Exception: log_data = []
    return {"files": log_data, "total_accounts": len(pm), "has_summary": SUMMARY_FILE.exists()}

@app.post("/api/dev/history/upload")
async def history_upload(request: Request, file: UploadFile = File(...), _=Depends(require_dev_token)):
    if not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(400, "只接受 .xlsx / .xls 格式")
    dest = HISTORY_DIR / file.filename
    dest.write_bytes(await file.read())
    total = rebuild_summary()
    log_data = []
    if UPLOAD_LOG.exists():
        try: log_data = json.loads(UPLOAD_LOG.read_text(encoding="utf-8"))
        except Exception: log_data = []
    log_data.insert(0, {"filename": file.filename, "uploaded_at": datetime.now().strftime("%Y-%m-%d %H:%M")})
    UPLOAD_LOG.write_text(json.dumps(log_data, ensure_ascii=False, indent=2), encoding="utf-8")
    return {"ok": True, "total_accounts": total}

@app.delete("/api/dev/history/reset")
async def history_reset(request: Request, _=Depends(require_dev_token)):
    if SUMMARY_FILE.exists(): SUMMARY_FILE.unlink()
    if UPLOAD_LOG.exists(): UPLOAD_LOG.unlink()
    return {"ok": True, "total_accounts": 0}

@app.get("/api/dev/history/note")
async def get_history_note(request: Request, _=Depends(require_dev_token)):
    return {"note": HISTORY_NOTE.read_text(encoding="utf-8") if HISTORY_NOTE.exists() else ""}

@app.post("/api/dev/history/note")
async def save_history_note(request: Request, _=Depends(require_dev_token)):
    body = await request.json()
    HISTORY_NOTE.write_text(body.get("note", ""), encoding="utf-8")
    return {"ok": True}
